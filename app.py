import os
import re
import io
import json
import time
import requests
import openpyxl
import streamlit as st
import pandas as pd
from streamlit_option_menu import option_menu
from pathlib import Path
from datetime import date, datetime, timedelta, timezone
from collections import Counter
from openpyxl.styles import Font, PatternFill, Alignment
from onboarding_wrapper import OnboardingAutomation, STEPS
from task_store import (
    load_tasks, add_task, update_task, delete_task, reset_recurring_tasks,
    load_members, add_member, update_member, delete_member,
)

# ── Config ────────────────────────────────────────────────────────────────────
# Airtable credentials come from Streamlit secrets (Settings → Secrets) or the
# AIRTABLE_PAT environment variable.
#
# NEVER hardcode a token here. This repository is PUBLIC, and a committed secret
# stays retrievable from git history forever even after it is removed from HEAD.
#
# `AIRTABLE_TOKEN` is the name task_store.py and onboarding_automation.py already
# use, so an existing secret is picked up with no config change. The lowercase and
# AIRTABLE_PAT spellings are accepted as fallbacks.
#
# NOTE ON SCOPE: task_store.py only needs the Tasks Overflow base. This module also
# reads v1 and v1.3 (see BASE_IDS) and the FH Contact Pull Tracker, so the token must
# be scoped to those bases too — otherwise those sections 403 while Tasks works fine.
def _first_secret(*names: str) -> str:
    for n in names:
        try:
            v = st.secrets.get(n)
        except Exception:
            v = None
        if v:
            return str(v)
        v = os.environ.get(n)
        if v:
            return v
    return ""


TOKEN = _first_secret("AIRTABLE_TOKEN", "airtable_token", "AIRTABLE_PAT")
HEADERS = {"Authorization": f"Bearer {TOKEN}"}
AIRTABLE_READY = bool(TOKEN)
BASE_IDS = ["appbXFzZnhij88tnQ", "appoDQDrqyvyPsZTY"]


def airtable_secret_warning() -> bool:
    """Render a warning if the Airtable token is missing. Returns True if missing."""
    if AIRTABLE_READY:
        return False
    st.warning(
        "**Airtable token not configured.** Add this to Streamlit Cloud secrets "
        "(Settings → Secrets), then rerun:\n\n"
        "`airtable_token = \"pat...\"`\n\n"
        "Locally, add the same line to `.streamlit/secrets.toml` (already gitignored). "
        "Airtable-backed sections will be empty until this is set."
    )
    return True

TARGET = re.compile(r"^\+1\d{10}$")
PLACEHOLDER_PATTERNS = [
    re.compile(r"\{[^}]+\}"),
    re.compile(r"\[[A-Z][^\]]+\]"),
    re.compile(r"<[A-Z][^>]+>"),
    re.compile(r"\{\{[^}]+\}\}"),
]

# ── Helpers ───────────────────────────────────────────────────────────────────
def get_base_name(base_id):
    r = requests.get("https://api.airtable.com/v0/meta/bases", headers=HEADERS)
    r.raise_for_status()
    for b in r.json().get("bases", []):
        if b["id"] == base_id:
            return b["name"]
    return base_id

def categorize_phone(value):
    if not value or not str(value).strip():
        return "Empty"
    v = str(value).strip()
    digits = re.sub(r"\D", "", v)
    if TARGET.match(v):
        return "OK"
    if len(digits) == 11 and digits.startswith("1"):
        return "Has digits but wrong format"
    if len(digits) == 10:
        return "Missing country code (+1)"
    if len(digits) > 11:
        return "Too many digits"
    if len(digits) < 10:
        return "Too few digits"
    return "Non-standard format"

def fix_phone_number(value):
    """Return reformatted E.164 number, or None if not auto-fixable."""
    digits = re.sub(r"\D", "", str(value).strip())
    if len(digits) == 10:
        return f"+1{digits}"
    if len(digits) == 11 and digits.startswith("1"):
        return f"+{digits}"
    return None

def has_placeholder(text):
    return any(p.search(text) for p in PLACEHOLDER_PATTERNS)

def extract_tokens(content):
    """Return a comma-separated string of every placeholder token found in content."""
    found = []
    for p in PLACEHOLDER_PATTERNS:
        found.extend(p.findall(str(content)))
    return ", ".join(sorted(set(found))) if found else ""

def categorize_message(content):
    if not content or not str(content).strip():
        return "Empty"
    if len(str(content).strip()) < 20:
        return "Too short"
    if has_placeholder(str(content)):
        return "Unfilled placeholder"
    return "OK"

def patch_phone_records(base_id, rows):
    """
    Patch Contact Cell for a list of records in batches of 10 (Airtable limit).
    rows: list of dicts with 'record_id' and 'fixed_value' keys.
    Returns (success_count, error_record_ids).
    """
    url = f"https://api.airtable.com/v0/{base_id}/Contact%20List"
    success, errors = 0, []
    for i in range(0, len(rows), 10):
        batch = rows[i:i + 10]
        payload = {
            "records": [
                {"id": r["record_id"], "fields": {"Contact Cell": r["Fixed Value"]}}
                for r in batch
            ]
        }
        try:
            resp = requests.patch(url, headers=HEADERS, json=payload)
            resp.raise_for_status()
            success += len(batch)
        except Exception:
            errors.extend([r["record_id"] for r in batch])
        time.sleep(0.22)   # ~4.5 req/sec — safely under Airtable's 5 req/sec limit
    return success, errors

def revert_phone_records(base_id, revert_rows):
    """
    Restore Contact Cell to original values.
    revert_rows: list of dicts with 'record_id' and 'original_value' keys.
    Returns (success_count, error_record_ids).
    """
    url = f"https://api.airtable.com/v0/{base_id}/Contact%20List"
    success, errors = 0, []
    for i in range(0, len(revert_rows), 10):
        batch = revert_rows[i:i + 10]
        payload = {
            "records": [
                {"id": r["record_id"], "fields": {"Contact Cell": r["original_value"]}}
                for r in batch
            ]
        }
        try:
            resp = requests.patch(url, headers=HEADERS, json=payload)
            resp.raise_for_status()
            success += len(batch)
        except Exception:
            errors.extend([r["record_id"] for r in batch])
        time.sleep(0.22)
    return success, errors

def fetch_records(base_id, table, fields, filter_formula=None, cell_format=None):
    records, offset = [], None
    url = f"https://api.airtable.com/v0/{base_id}/{requests.utils.quote(table)}"
    while True:
        params = {"pageSize": 100, "fields[]": fields}
        if offset:
            params["offset"] = offset
        if filter_formula:
            params["filterByFormula"] = filter_formula
        if cell_format:
            params["cellFormat"] = cell_format
        r = requests.get(url, headers=HEADERS, params=params)
        r.raise_for_status()
        data = r.json()
        records.extend(data.get("records", []))
        offset = data.get("offset")
        if not offset:
            break
    return records

def run_phone_audit(base_id, base_name):
    records = fetch_records(base_id, "Contact List",
                            ["Contact Cell", "Contact Full Name:", "Funeral Home Name"])
    rows = []
    for rec in records:
        f = rec.get("fields", {})
        val = f.get("Contact Cell", "")
        name = f.get("Contact Full Name:", "(unknown)")
        fh_raw = f.get("Funeral Home Name", [])
        fh = fh_raw[0] if isinstance(fh_raw, list) and fh_raw else str(fh_raw) if fh_raw else "(unknown)"
        cat = categorize_phone(val)
        rows.append({
            "Base": base_name,
            "Funeral Home": fh,
            "Contact Full Name": name,
            "Record ID": rec["id"],
            "Current Value": str(val) if val else "(empty)",
            "Issue": cat,
        })
    return pd.DataFrame(rows)

def run_message_audit(base_id, base_name):
    records = fetch_records(
        base_id, "Messages",
        ["Direction", "Message Content", "Message Type",
         "Contact Full Name: (from Contact Cell)",
         "Funeral Home: (from Contact Cell)"],
        filter_formula='FIND("outbound", LOWER({Direction})) > 0',
        cell_format="string"
    )
    rows = []
    for rec in records:
        f = rec.get("fields", {})
        content = f.get("Message Content", "")
        name_raw = f.get("Contact Full Name: (from Contact Cell)", "")
        name = name_raw[0] if isinstance(name_raw, list) and name_raw else str(name_raw) if name_raw else "(unknown)"
        fh_raw = f.get("Funeral Home: (from Contact Cell)", "")
        fh = fh_raw if isinstance(fh_raw, str) and fh_raw else (fh_raw[0] if isinstance(fh_raw, list) and fh_raw else "(unknown)")
        cat = categorize_message(content)
        rows.append({
            "Base": base_name,
            "Funeral Home": fh,
            "Contact Full Name": name,
            "Record ID": rec["id"],
            "Direction": f.get("Direction", ""),
            "Message Type": f.get("Message Type", "(unknown)"),
            "Issue": cat,
            "Content (first 200 chars)": str(content)[:200] if content else "(empty)",
        })
    return pd.DataFrame(rows)

def build_excel(results_dict):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    HEADER_FILL = PatternFill("solid", start_color="1a2b4a")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    ISSUE_COLORS = {
        "Has digits but wrong format": "DDEBF7",
        "Missing country code (+1)": "FFF2CC",
        "Too many digits": "FCE4D6",
        "Too few digits": "FCE4D6",
        "Non-standard format": "EAD1DC",
        "Empty": "E2EFDA",
        "Unfilled placeholder": "FFF2CC",
        "Too short": "FCE4D6",
    }
    for sheet_name, df in results_dict.items():
        safe = sheet_name[:31]
        ws = wb.create_sheet(safe)
        for c, col in enumerate(df.columns, 1):
            cell = ws.cell(1, c, col)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        for r, row in enumerate(df.itertuples(index=False), 2):
            issue = row.Issue if hasattr(row, "Issue") else ""
            for c, val in enumerate(row, 1):
                cell = ws.cell(r, c, val)
                if issue in ISSUE_COLORS:
                    cell.fill = PatternFill("solid", start_color=ISSUE_COLORS[issue])
        for c, col in enumerate(df.columns, 1):
            max_len = max(len(str(col)), df.iloc[:, c-1].astype(str).str.len().max() if len(df) else 0)
            ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = min(max_len + 4, 50)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(df.columns))}1"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ── Page Config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Aftercare Texting Audit — Parting Pro",
    layout="wide",
    page_icon="https://partingpro.com/wp-content/uploads/2024/07/partingpro-logo.png"
)

# ── Dark-mode session state (read at top so CSS picks it up on every rerun) ──
if "dark_mode" not in st.session_state:
    st.session_state["dark_mode"] = False
_DARK = bool(st.session_state["dark_mode"])

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
@import url('https://api.fontshare.com/v2/css?f[]=satoshi@400,500,700,900&display=swap');

/* Parting Pro brand palette — pulled from partingpro.com's theme (style.css --*-color vars) */
:root {
    --pp-dark: #0B0F31;
    --pp-dark-mid: #142352;
    --pp-text: #333864;
    --pp-accent: #0D6DA3;
    --pp-accent-hover: #1B7BB1;
    --pp-blue: #166AE8;
    --pp-accent-tint: #E6F2F8;
    --pp-green: #1B9E6B;
    --pp-error: #FB3D3D;
    --pp-orange: #E26514;
    --pp-warning: #f9c127;
    --pp-border: #E3E5E9;
    --pp-bg: #F3F7FA;
}

html, body, [class*="css"] {
    font-family: 'Satoshi', 'Inter', sans-serif;
}

/* Hide default streamlit header */
#MainMenu, footer, header { visibility: hidden; }

.stApp { background: var(--pp-bg); }

/* ── Hero ── */
.hero {
    background: linear-gradient(135deg, var(--pp-dark) 0%, var(--pp-dark-mid) 55%, var(--pp-accent) 140%);
    border-radius: 16px;
    padding: 48px 56px;
    margin-bottom: 32px;
    display: flex;
    align-items: center;
    justify-content: space-between;
    position: relative;
    overflow: hidden;
    box-shadow: 0 8px 32px rgba(11,15,49,0.22);
}
.hero::before {
    content: '';
    position: absolute;
    top: -60px; right: -60px;
    width: 280px; height: 280px;
    background: rgba(255,255,255,0.04);
    border-radius: 50%;
}
.hero::after {
    content: '';
    position: absolute;
    bottom: -80px; right: 120px;
    width: 200px; height: 200px;
    background: rgba(255,255,255,0.03);
    border-radius: 50%;
}
.hero-left { z-index: 1; }
.hero-logo { height: 36px; margin-bottom: 20px; filter: brightness(0) invert(1); }
.hero-title {
    font-size: 30px;
    font-weight: 700;
    color: #ffffff;
    margin: 0 0 8px 0;
    line-height: 1.2;
    letter-spacing: -0.5px;
}
.hero-subtitle {
    font-size: 15px;
    color: rgba(255,255,255,0.65);
    margin: 0;
    font-weight: 400;
}
.hero-badge {
    background: rgba(255,255,255,0.1);
    border: 1px solid rgba(255,255,255,0.2);
    border-radius: 8px;
    padding: 8px 16px;
    color: rgba(255,255,255,0.85);
    font-size: 12px;
    font-weight: 500;
    z-index: 1;
    backdrop-filter: blur(8px);
}

/* ── Cards ── */
.card {
    background: white;
    border-radius: 12px;
    padding: 24px;
    border: 1px solid #e4e7ef;
    box-shadow: 0 1px 4px rgba(0,0,0,0.05);
    margin-bottom: 16px;
}
.card-title {
    font-size: 15px;
    font-weight: 600;
    color: var(--pp-text);
    margin-bottom: 16px;
    display: flex;
    align-items: center;
    gap: 8px;
}

/* ── Metric Cards ── */
.metrics-row { display: flex; gap: 16px; margin-bottom: 20px; }
.metric {
    flex: 1;
    background: white;
    border-radius: 12px;
    padding: 20px 24px;
    border: 1px solid #e4e7ef;
    box-shadow: 0 1px 4px rgba(0,0,0,0.04);
}
.metric .m-label {
    font-size: 11px;
    font-weight: 600;
    color: #4a5568;
    text-transform: uppercase;
    letter-spacing: 0.06em;
    margin-bottom: 8px;
}
.metric .m-value {
    font-size: 36px;
    font-weight: 700;
    color: var(--pp-text);
    line-height: 1;
}
.metric .m-sub {
    font-size: 12px;
    color: #4a5568;
    margin-top: 4px;
}
.metric.green .m-value { color: var(--pp-green); }
.metric.red .m-value { color: var(--pp-error); }
.metric.blue .m-value { color: var(--pp-accent); }

/* ── Section Headers ── */
.section-wrap {
    background: white;
    border-radius: 16px;
    padding: 28px 32px;
    border: 1px solid #e4e7ef;
    box-shadow: 0 1px 6px rgba(0,0,0,0.04);
    margin-bottom: 24px;
}
.section-head {
    display: flex;
    align-items: center;
    gap: 12px;
    margin-bottom: 20px;
    padding-bottom: 16px;
    border-bottom: 1px solid #f0f2f7;
}
.section-icon {
    width: 40px; height: 40px;
    background: var(--pp-accent-tint);
    border-radius: 10px;
    display: flex; align-items: center; justify-content: center;
    font-size: 20px;
}
.section-head-text h3 {
    font-size: 17px;
    font-weight: 650;
    color: var(--pp-text);
    margin: 0 0 2px 0;
}
.section-head-text p {
    font-size: 13px;
    color: #4a5568;
    margin: 0;
}

/* ── Base Tag ── */
.base-tag {
    display: inline-flex;
    align-items: center;
    gap: 6px;
    background: var(--pp-accent-tint);
    color: var(--pp-accent);
    border-radius: 6px;
    padding: 4px 12px;
    font-size: 12px;
    font-weight: 600;
    margin-bottom: 16px;
}

/* ── Issue Badge ── */
.issue-ok { color: var(--pp-green); font-weight: 600; }
.issue-warn { color: var(--pp-orange); font-weight: 600; }
.issue-error { color: var(--pp-error); font-weight: 600; }

/* ══ GLOBAL TEXT VISIBILITY — main content area only ══════════════════ */

/* Every p, span, label, div text in the main block */
section[data-testid="stMain"] p,
section[data-testid="stMain"] span,
section[data-testid="stMain"] li,
section[data-testid="stMain"] strong,
section[data-testid="stMain"] em,
section[data-testid="stMain"] h1,
section[data-testid="stMain"] h2,
section[data-testid="stMain"] h3,
section[data-testid="stMain"] h4,
section[data-testid="stMain"] h5 {
    color: var(--pp-text) !important;
}

/* Checkbox label text */
section[data-testid="stMain"] .stCheckbox label,
section[data-testid="stMain"] .stCheckbox label p,
section[data-testid="stMain"] [data-testid="stCheckbox"] label {
    color: var(--pp-text) !important;
    font-weight: 500 !important;
    font-size: 14px !important;
}

/* Number input label + field */
section[data-testid="stMain"] .stNumberInput label,
section[data-testid="stMain"] .stNumberInput label p,
section[data-testid="stMain"] [data-testid="stNumberInput"] label {
    color: var(--pp-text) !important;
    font-weight: 500 !important;
    font-size: 14px !important;
}
section[data-testid="stMain"] .stNumberInput input {
    color: var(--pp-text) !important;
    background: #ffffff !important;
    border: 1px solid #c8cdd8 !important;
}

/* Spinner / loading text */
section[data-testid="stMain"] [data-testid="stSpinner"] p,
section[data-testid="stMain"] [data-testid="stSpinner"] span,
section[data-testid="stMain"] [data-testid="stSpinnerContainer"] p,
section[data-testid="stMain"] .stSpinner p {
    color: var(--pp-text) !important;
    font-weight: 500 !important;
}

/* Alert / banner body text */
section[data-testid="stMain"] [data-testid="stAlert"] p,
section[data-testid="stMain"] .stAlert p {
    font-weight: 500 !important;
}

/* Bar chart axis labels */
section[data-testid="stMain"] .vega-embed text,
section[data-testid="stMain"] .vega-embed .mark-text text {
    fill: var(--pp-text) !important;
}

/* Hero text must stay white — earlier `color: inherit` was leaking dark body color through */
.hero, .hero h1, .hero h2, .hero h3, .hero p, .hero div, .hero span { color: #ffffff !important; }
.hero .hero-subtitle { color: rgba(255,255,255,0.75) !important; }
.hero .hero-badge { color: rgba(255,255,255,0.9) !important; }
section[data-testid="stSidebar"] * { color: rgba(255,255,255,0.85) !important; }

/* ── Sidebar ── */
section[data-testid="stSidebar"] {
    background: var(--pp-dark) !important;
}
section[data-testid="stSidebar"] * {
    color: rgba(255,255,255,0.85) !important;
}
section[data-testid="stSidebar"] .stButton > button {
    background: rgba(255,255,255,0.1) !important;
    color: white !important;
    border: 1px solid rgba(255,255,255,0.2) !important;
    border-radius: 8px !important;
    font-weight: 500 !important;
    transition: background 0.2s !important;
    width: 100%;
}
section[data-testid="stSidebar"] .stButton > button:hover {
    background: rgba(255,255,255,0.2) !important;
}
section[data-testid="stSidebar"] hr {
    border-color: rgba(255,255,255,0.1) !important;
}

/* ── Download Button ── */
div[data-testid="stDownloadButton"] > button {
    background: var(--pp-dark) !important;
    color: white !important;
    border: none !important;
    border-radius: 8px !important;
    padding: 10px 20px !important;
    font-weight: 500 !important;
}
div[data-testid="stDownloadButton"] > button:hover {
    background: var(--pp-accent-hover) !important;
}

/* ── Dataframe ── */
.stDataFrame { border-radius: 10px; overflow: hidden; }

/* ── Divider ── */
hr { border-color: #e4e7ef !important; margin: 24px 0 !important; }

/* ── Task Tracker — Priority Pills ── */
.pill-p1 {
    display: inline-block;
    background: #fde8e8; color: #c0392b;
    border: 1px solid #f5c6c6;
    border-radius: 20px; padding: 2px 10px;
    font-size: 11px; font-weight: 700; letter-spacing: 0.04em;
}
.pill-p2 {
    display: inline-block;
    background: #fff0e0; color: #c47f00;
    border: 1px solid #f5d9a0;
    border-radius: 20px; padding: 2px 10px;
    font-size: 11px; font-weight: 700; letter-spacing: 0.04em;
}
.pill-p3 {
    display: inline-block;
    background: #f0f2f7; color: #4a5568;
    border: 1px solid #d0d5e0;
    border-radius: 20px; padding: 2px 10px;
    font-size: 11px; font-weight: 700; letter-spacing: 0.04em;
}
.type-badge {
    display: inline-block;
    background: var(--pp-accent-tint); color: var(--pp-accent);
    border-radius: 6px; padding: 2px 8px;
    font-size: 11px; font-weight: 600;
    text-transform: capitalize;
}
.task-title { font-size: 14px; font-weight: 600; color: var(--pp-text); }
.task-title-done { font-size: 14px; font-weight: 500; color: #9aa5b4; text-decoration: line-through; }
.task-desc { font-size: 12px; color: #6b7a94; margin-top: 2px; }
.overdue { color: var(--pp-error) !important; font-weight: 600 !important; }
.due-ok { color: #4a5568; }
</style>
""", unsafe_allow_html=True)

# ── Dark-mode CSS overrides ───────────────────────────────────────────────────
# Injected only when the sidebar toggle is on. We override the light theme
# rather than rewriting it so the light theme stays the default and unchanged.
if _DARK:
    st.markdown("""
    <style>
    /* Page background + global text */
    .stApp { background: #0d1117 !important; }
    section[data-testid="stMain"] { background: #0d1117 !important; }

    /* Cards, sections, panels — dark surface */
    .card, .section-wrap, .metric, .empty {
        background: #161b22 !important;
        border-color: #2a3142 !important;
        box-shadow: 0 1px 6px rgba(0,0,0,0.4) !important;
    }
    .section-head { border-bottom-color: #2a3142 !important; }
    hr { border-color: #2a3142 !important; }

    /* Primary text — flip dark navy to soft light gray */
    section[data-testid="stMain"] p,
    section[data-testid="stMain"] span,
    section[data-testid="stMain"] li,
    section[data-testid="stMain"] strong,
    section[data-testid="stMain"] em,
    section[data-testid="stMain"] h1,
    section[data-testid="stMain"] h2,
    section[data-testid="stMain"] h3,
    section[data-testid="stMain"] h4,
    section[data-testid="stMain"] h5,
    .card-title, .section-head-text h3, .metric .m-value,
    .task-title {
        color: #e4e7ef !important;
    }
    .section-head-text p, .metric .m-label, .metric .m-sub,
    .task-desc, .due-ok {
        color: #9aa5b4 !important;
    }

    /* Hero still has its own background gradient — keep text crisp white */
    .hero, .hero * { color: #ffffff !important; }
    .hero .hero-subtitle { color: rgba(255,255,255,0.75) !important; }

    /* Inputs / selects / textareas */
    section[data-testid="stMain"] input,
    section[data-testid="stMain"] textarea,
    section[data-testid="stMain"] select,
    section[data-testid="stMain"] .stNumberInput input,
    section[data-testid="stMain"] .stDateInput input,
    section[data-testid="stMain"] [data-baseweb="input"] input,
    section[data-testid="stMain"] [data-baseweb="select"] > div {
        background: #161b22 !important;
        color: #e4e7ef !important;
        border-color: #2a3142 !important;
    }

    /* Tab bar */
    section[data-testid="stMain"] [data-baseweb="tab-list"] {
        background: #0d1117 !important;
        border-bottom-color: #2a3142 !important;
    }
    section[data-testid="stMain"] [data-baseweb="tab"] { color: #9aa5b4 !important; }
    section[data-testid="stMain"] [data-baseweb="tab"][aria-selected="true"] {
        color: #e4e7ef !important;
    }

    /* Expanders */
    section[data-testid="stMain"] [data-testid="stExpander"] {
        background: #161b22 !important;
        border: 1px solid #2a3142 !important;
        border-radius: 8px !important;
    }
    section[data-testid="stMain"] [data-testid="stExpander"] summary { color: #e4e7ef !important; }

    /* Code / pre blocks */
    section[data-testid="stMain"] code,
    section[data-testid="stMain"] pre {
        background: #0d1117 !important;
        color: #e4e7ef !important;
        border: 1px solid #2a3142 !important;
    }

    /* Dataframes */
    section[data-testid="stMain"] .stDataFrame { background: #161b22 !important; }
    section[data-testid="stMain"] [data-testid="stDataFrame"] * { color: #e4e7ef !important; }

    /* Alerts — keep their semantic background but darken slightly */
    section[data-testid="stMain"] [data-testid="stAlert"] {
        background: #161b22 !important;
        border-left-width: 3px !important;
    }

    /* Sidebar already dark — just deepen it a touch */
    section[data-testid="stSidebar"] { background: #0a0e14 !important; }

    /* Buttons in the main area */
    section[data-testid="stMain"] .stButton > button {
        background: #1f2937 !important;
        color: #e4e7ef !important;
        border-color: #2a3142 !important;
    }
    section[data-testid="stMain"] .stButton > button:hover {
        background: #2a3441 !important;
        border-color: var(--pp-accent) !important;
    }

    /* Pills + badges — invert the light pastel backgrounds to fit dark mode */
    .pill-p3 { background: #1f2937 !important; color: #c8cdd8 !important; border-color: #2a3142 !important; }
    .type-badge { background: #1c2a4a !important; color: #93b3ff !important; }
    .base-tag { background: #1c2a4a !important; color: #93b3ff !important; }
    </style>
    """, unsafe_allow_html=True)

# ── Task Tracker — session state ──────────────────────────────────────────────
for _k in ("editing_task_id", "deleting_task_id"):
    if _k not in st.session_state:
        st.session_state[_k] = None

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div style="text-align:center; padding: 16px 0 8px 0;">
        <img src="https://partingpro.com/wp-content/uploads/2024/07/partingpro-logo_white.png"
             style="height:28px; filter: brightness(0) invert(1);" />
    </div>
    """, unsafe_allow_html=True)

    # ── Page navigation ───────────────────────────────────────────────────
    # Add new pages here as they're ready — e.g. "Texting Audit": render_texting_audit
    PAGE_RENDERERS = {
        "Onboarding": None,  # bound to render_onboarding() further down the file
        "Zap Audit":  None,  # bound to render_zap_audit() further down the file
        "Tasks":      None,  # bound to render_tasks() further down the file
        "Dashboard":  None,  # bound to render_dashboard() further down the file
    }
    selected_page = option_menu(
        menu_title=None,
        options=list(PAGE_RENDERERS.keys()),
        icons=["rocket-takeoff", "lightning-charge", "check2-square", "bar-chart-line"],
        default_index=0,
        styles={
            "container": {"padding": "0", "background-color": "transparent"},
            "icon": {"color": "#333864", "font-size": "15px"},
            "nav-link": {
                "font-size": "14px",
                "text-align": "left",
                "margin": "2px 0",
                "color": "#333864",
                "background-color": "transparent",
                "border-radius": "8px",
                "--hover-color": "rgba(11,15,49,0.08)",
            },
            "nav-link-selected": {
                "background-color": "#0D6DA3",
                "color": "#ffffff",
                "font-weight": "600",
            },
        },
    )

    # Dark-mode toggle — bound to session_state["dark_mode"] via key; flip triggers rerun
    st.toggle("🌙  Dark mode", key="dark_mode", help="Switch the main content to a dark theme")
    st.markdown("---")
    st.markdown("<div style='font-size:11px; font-weight:600; text-transform:uppercase; letter-spacing:0.08em; opacity:0.5; margin-bottom:8px;'>Connected Bases</div>", unsafe_allow_html=True)
    for b in BASE_IDS:
        st.markdown(f"<div style='font-size:12px; opacity:0.7; padding: 4px 0;'>• {b}</div>", unsafe_allow_html=True)
    st.markdown("---")
    st.markdown("<div style='font-size:11px; opacity:0.4; text-align:center;'>Parting Pro Internal · 2025</div>", unsafe_allow_html=True)
    # "Add Task" quick-form — only shown when the Tasks page is enabled in
    # PAGE_RENDERERS above (hidden for now since only Onboarding is exposed).
    if selected_page == "Tasks":
        st.markdown("---")
        st.markdown("<div style='font-size:11px; font-weight:600; text-transform:uppercase; letter-spacing:0.08em; opacity:0.5; margin-bottom:12px;'>Add Task</div>", unsafe_allow_html=True)
        _members_for_form = [m for m in load_members() if m.get("active")]
        _name_to_id = {m["name"]: m["id"] for m in _members_for_form}
        with st.form("sidebar_add_task", clear_on_submit=True):
            _title = st.text_input("Title *", placeholder="What needs to be done?")
            _desc  = st.text_area("Description", placeholder="Optional…", height=60)
            _type  = st.selectbox("Type", ["daily", "weekly", "monthly", "one-off"])
            _pri   = st.selectbox("Priority", ["P1", "P2", "P3"], index=1)
            _due   = st.date_input("Due Date (optional)", value=None) if _type == "one-off" else None
            _assignees = st.multiselect("Assign to", options=list(_name_to_id.keys()),
                                        help="Pick from active team members. Manage the roster from the Tasks tab.")
            _sub   = st.form_submit_button("➕ Add Task", use_container_width=True)
        if _sub:
            if _title.strip():
                add_task({"title": _title.strip(), "description": _desc.strip(),
                          "type": _type, "priority": _pri,
                          "due_date": str(_due) if _due else None,
                          "assignee_ids": [_name_to_id[n] for n in _assignees]})
                st.rerun()
            else:
                st.warning("Title required.")

# Recurring-task reset only runs when the Tasks page is actually being viewed —
# previously this ran on every rerun of the whole app regardless of page, which
# polled the Task Tracker base constantly and burned through its monthly
# Airtable API quota (the billing-limit error banner).
if selected_page == "Tasks":
    reset_recurring_tasks()

# ── Hero Section ──────────────────────────────────────────────────────────────
# Copy per page — was a single static "Aftercare Texting" banner shown on every
# page regardless of which one was selected.
_HERO_COPY = {
    "Onboarding": ("Onboarding Automation", "Walk a new funeral home through setup, step by step", "Airtable Connected"),
    "Zap Audit":  ("Zap Audit — Live", "Every zap run logs to Supabase; monitor status in real time", "Supabase Connected"),
    "Tasks":      ("Team Tasks", "Track daily, weekly, monthly, and one-off work across the team", "Airtable Connected"),
    "Dashboard":  ("Task Dashboard", "Filter and explore team task activity", "Airtable Connected"),
}
_hero_title, _hero_subtitle, _hero_badge = _HERO_COPY.get(
    selected_page, ("Parting Pro", "Internal tool", "Internal Tool")
)
st.markdown(f"""
<div class="hero">
    <div class="hero-left">
        <img class="hero-logo" src="https://partingpro.com/wp-content/uploads/2024/07/partingpro-logo_white.png" />
        <div class="hero-title">{_hero_title}</div>
        <div class="hero-subtitle">{_hero_subtitle}</div>
    </div>
    <div class="hero-badge">🔒 Internal Tool &nbsp;·&nbsp; {_hero_badge}</div>
</div>
""", unsafe_allow_html=True)

# ── Pages ─────────────────────────────────────────────────────────────────────
# Each page below is a plain function; the sidebar nav (top of file) picks which
# one runs. Only "Onboarding" is wired into PAGE_RENDERERS for now — the others
# stay ready to re-enable (e.g. add "Dashboard": render_zap_audit later).

def render_texting_audit():
    # ── Audit Controls (Texting tab) ──────────────────────────────────────────────
    _phone_col, _msg_col = st.columns(2)
    with _phone_col:
        run_phones = st.button("📞  Run Phone Audit", use_container_width=True, key="run_phones_btn")
    with _msg_col:
        run_messages = st.button("💬  Run Message Audit", use_container_width=True, key="run_messages_btn")
    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

    # ── Phone Audit ───────────────────────────────────────────────────────────────
    st.markdown("""
    <div class="section-wrap">
        <div class="section-head">
            <div class="section-icon">📞</div>
            <div class="section-head-text">
                <h3>Step 1 — Phone Number Audit</h3>
                <p>Validates Contact Cell format against E.164 standard (+1XXXXXXXXXX)</p>
            </div>
        </div>
    """, unsafe_allow_html=True)

    if run_phones:
        for base_id in BASE_IDS:
            with st.spinner(f"Fetching records from {base_id}..."):
                base_name = get_base_name(base_id)
                df = run_phone_audit(base_id, base_name)
                st.session_state[f"phone_{base_id}"] = df
                st.session_state[f"phone_name_{base_id}"] = base_name
        st.success("✅ Phone audit complete for both bases!")

    for base_id in BASE_IDS:
        if f"phone_{base_id}" in st.session_state:
            df = st.session_state[f"phone_{base_id}"]
            base_name = st.session_state[f"phone_name_{base_id}"]

            total = len(df)
            ok = len(df[df["Issue"] == "OK"])
            flagged = len(df[df["Issue"] != "OK"])
            pass_rate = round((ok / total * 100), 1) if total else 0

            st.markdown(f'<div class="base-tag">🏢 {base_name}</div>', unsafe_allow_html=True)
            st.markdown(f"""
            <div class="metrics-row">
                <div class="metric blue">
                    <div class="m-label">Total Records</div>
                    <div class="m-value">{total:,}</div>
                    <div class="m-sub">Contact List</div>
                </div>
                <div class="metric green">
                    <div class="m-label">✅ Passing</div>
                    <div class="m-value">{ok:,}</div>
                    <div class="m-sub">{pass_rate}% pass rate</div>
                </div>
                <div class="metric red">
                    <div class="m-label">⚠️ Flagged</div>
                    <div class="m-value">{flagged:,}</div>
                    <div class="m-sub">Need attention</div>
                </div>
            </div>
            """, unsafe_allow_html=True)

            issue_counts = df[df["Issue"] != "OK"]["Issue"].value_counts().reset_index()
            issue_counts.columns = ["Issue", "Count"]
            if not issue_counts.empty:
                st.markdown("**Issue Breakdown**")
                st.bar_chart(issue_counts.set_index("Issue"), color="#333864")

            flagged_df = df[df["Issue"] != "OK"]
            if not flagged_df.empty:
                st.markdown(f"**Flagged Records — {len(flagged_df)} total**")
                st.dataframe(flagged_df, use_container_width=True, hide_index=True)

            # ── Auto-Fix Section ──────────────────────────────────────────────
            FIXABLE_ISSUES = {"Missing country code (+1)", "Has digits but wrong format"}
            fixable_rows = []
            for _, row in df[df["Issue"].isin(FIXABLE_ISSUES)].iterrows():
                fixed = fix_phone_number(row["Current Value"])
                if fixed:
                    fixable_rows.append({
                        "record_id": row["Record ID"],
                        "Contact Full Name": row["Contact Full Name"],
                        "Funeral Home": row["Funeral Home"],
                        "Current Value": row["Current Value"],
                        "Fixed Value": fixed,
                        "Issue": row["Issue"],
                    })

            if fixable_rows:
                fix_df = pd.DataFrame(fixable_rows)

                # Track which record IDs have already been patched this session
                applied_key = f"fix_applied_{base_id}"
                if applied_key not in st.session_state:
                    st.session_state[applied_key] = set()

                pending = [r for r in fixable_rows
                           if r["record_id"] not in st.session_state[applied_key]]
                n_done = len(fixable_rows) - len(pending)

                st.markdown(f"**🔧 {len(fix_df)} number(s) can be auto-fixed**")
                st.dataframe(
                    fix_df[["Contact Full Name", "Funeral Home", "Current Value", "Fixed Value", "Issue"]],
                    use_container_width=True, hide_index=True
                )

                if n_done:
                    st.success(f"✅ {n_done} of {len(fixable_rows)} record(s) fixed so far this session.")

                if pending:
                    confirmed = st.checkbox(
                        f"I've reviewed the changes above and want to apply them to {base_name}",
                        key=f"confirm_fix_{base_id}"
                    )
                    if confirmed:
                        max_test = min(10, len(pending))
                        test_n = int(st.number_input(
                            f"How many records to patch first? (max 10 for a safe test run)",
                            min_value=1, max_value=max_test, value=min(3, max_test),
                            key=f"test_n_{base_id}"
                        ))

                        if n_done == 0:
                            # No test run yet — only offer the test button
                            if st.button(f"🧪 Test fix ({test_n} record(s))", key=f"test_fix_{base_id}"):
                                with st.spinner(f"Patching {test_n} record(s) in Airtable…"):
                                    ok, errs = patch_phone_records(base_id, pending[:test_n])
                                for r in pending[:ok]:
                                    st.session_state[applied_key].add(r["record_id"])
                                if errs:
                                    st.warning(f"Fixed {ok}/{test_n}. ⚠️ {len(errs)} failed — try again.")
                                else:
                                    st.success(f"✅ Test passed — {ok} record(s) fixed. "
                                               f"Check Airtable to confirm, then apply the rest below.")
                                st.rerun()
                        else:
                            # Test already ran — offer both another test batch and apply-all
                            col1, col2 = st.columns(2)
                            with col1:
                                if st.button(f"🧪 Test another {test_n} record(s)",
                                             key=f"test_fix_{base_id}"):
                                    with st.spinner(f"Patching {test_n} record(s)…"):
                                        ok, errs = patch_phone_records(base_id, pending[:test_n])
                                    for r in pending[:ok]:
                                        st.session_state[applied_key].add(r["record_id"])
                                    if errs:
                                        st.warning(f"Fixed {ok}/{test_n}. ⚠️ {len(errs)} failed.")
                                    else:
                                        st.success(f"✅ Fixed {ok} more. "
                                                   f"{len(pending) - ok} remaining.")
                                    st.rerun()
                            with col2:
                                if st.button(f"✅ Apply all {len(pending)} remaining",
                                             key=f"apply_all_{base_id}"):
                                    with st.spinner(f"Patching {len(pending)} record(s)…"):
                                        ok, errs = patch_phone_records(base_id, pending)
                                    for r in pending[:ok]:
                                        st.session_state[applied_key].add(r["record_id"])
                                    if errs:
                                        st.warning(f"Fixed {ok}. ⚠️ {len(errs)} failed — re-run audit to retry.")
                                    else:
                                        st.success(f"✅ All done! Fixed {ok} records in {base_name}.")
                                    st.rerun()
                else:
                    # Every fixable record has been patched
                    st.success(f"✅ All {len(fixable_rows)} numbers in {base_name} are fixed!")
                    if st.button("🔄 Re-run audit to confirm", key=f"clear_{base_id}"):
                        del st.session_state[f"phone_{base_id}"]
                        del st.session_state[f"phone_name_{base_id}"]
                        if applied_key in st.session_state:
                            del st.session_state[applied_key]
                        st.rerun()

            elif flagged > 0:
                st.info("ℹ️ No auto-fixable numbers found — all flagged records need manual review in Airtable.")
            # ─────────────────────────────────────────────────────────────────
            st.markdown("---")

    st.markdown("</div>", unsafe_allow_html=True)

    if any(f"phone_{b}" in st.session_state for b in BASE_IDS):
        all_dfs = {
            f"{st.session_state[f'phone_name_{b}']} - Issues": st.session_state[f"phone_{b}"][st.session_state[f"phone_{b}"]["Issue"] != "OK"]
            for b in BASE_IDS if f"phone_{b}" in st.session_state
        }
        excel_buf = build_excel(all_dfs)
        st.download_button("⬇️ Download Phone Audit Report (.xlsx)",
                           excel_buf, "phone_audit_results.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)

    # ── Message Audit ─────────────────────────────────────────────────────────────
    st.markdown("""
    <div class="section-wrap">
        <div class="section-head">
            <div class="section-icon">💬</div>
            <div class="section-head-text">
                <h3>Step 2 — Message Content Audit</h3>
                <p>Scans outbound messages for unfilled placeholders, empty content, and short messages</p>
            </div>
        </div>
    """, unsafe_allow_html=True)

    if run_messages:
        for base_id in BASE_IDS:
            with st.spinner(f"Fetching outbound messages from {base_id}..."):
                base_name = get_base_name(base_id)
                df = run_message_audit(base_id, base_name)
                st.session_state[f"msg_{base_id}"] = df
                st.session_state[f"msg_name_{base_id}"] = base_name
        st.success("✅ Message audit complete for both bases!")

    for base_id in BASE_IDS:
        if f"msg_{base_id}" in st.session_state:
            df = st.session_state[f"msg_{base_id}"]
            base_name = st.session_state[f"msg_name_{base_id}"]

            # ── Test filter ───────────────────────────────────────────────
            excl_test = st.checkbox(
                "🔕 Exclude messages containing 'test'",
                value=True,
                key=f"excl_test_{base_id}"
            )
            df_view = (
                df[~df["Content (first 200 chars)"].str.contains("test", case=False, na=False)]
                if excl_test else df
            )

            total = len(df_view)
            ok = len(df_view[df_view["Issue"] == "OK"])
            flagged = len(df_view[df_view["Issue"] != "OK"])
            pass_rate = round((ok / total * 100), 1) if total else 0

            st.markdown(f'<div class="base-tag">🏢 {base_name}</div>', unsafe_allow_html=True)
            st.markdown(f"""
            <div class="metrics-row">
                <div class="metric blue">
                    <div class="m-label">Total Outbound</div>
                    <div class="m-value">{total:,}</div>
                    <div class="m-sub">Outbound messages</div>
                </div>
                <div class="metric green">
                    <div class="m-label">✅ Passing</div>
                    <div class="m-value">{ok:,}</div>
                    <div class="m-sub">{pass_rate}% pass rate</div>
                </div>
                <div class="metric red">
                    <div class="m-label">⚠️ Flagged</div>
                    <div class="m-value">{flagged:,}</div>
                    <div class="m-sub">Need attention</div>
                </div>
            </div>
            """, unsafe_allow_html=True)

            issue_counts = df_view[df_view["Issue"] != "OK"]["Issue"].value_counts().reset_index()
            issue_counts.columns = ["Issue", "Count"]
            if not issue_counts.empty:
                st.markdown("**Issue Breakdown**")
                st.bar_chart(issue_counts.set_index("Issue"), color="#333864")

            flagged_df = df_view[df_view["Issue"] != "OK"]
            if not flagged_df.empty:
                st.markdown(f"**Flagged Records — {len(flagged_df)} total**")
                st.dataframe(flagged_df, use_container_width=True, hide_index=True)

            # ── Placeholder Breakdown ─────────────────────────────────────
            ph_df = df_view[df_view["Issue"] == "Unfilled placeholder"].copy()
            if not ph_df.empty:
                ph_df["Bad Token(s)"] = ph_df["Content (first 200 chars)"].apply(extract_tokens)

                all_tokens = Counter()
                for content in ph_df["Content (first 200 chars)"]:
                    for p in PLACEHOLDER_PATTERNS:
                        for m in p.findall(str(content)):
                            all_tokens[m] += 1

                st.markdown("**📋 Unfilled Placeholder Breakdown**")
                col1, col2 = st.columns([1, 2])
                with col1:
                    st.markdown("**Token frequency**")
                    st.dataframe(
                        pd.DataFrame(all_tokens.most_common(), columns=["Bad Token", "Times Sent"]),
                        use_container_width=True, hide_index=True
                    )
                with col2:
                    st.markdown(f"**{len(ph_df)} affected message(s)**")
                    st.dataframe(
                        ph_df[["Contact Full Name", "Funeral Home",
                               "Bad Token(s)", "Content (first 200 chars)"]],
                        use_container_width=True, hide_index=True
                    )
                st.info(
                    "ℹ️ These messages were already sent with unfilled tokens. "
                    "The contacts above may need a follow-up message. "
                    "Fix the corresponding message templates to prevent future occurrences."
                )
            # ─────────────────────────────────────────────────────────────
            st.markdown("---")

    st.markdown("</div>", unsafe_allow_html=True)

    if any(f"msg_{b}" in st.session_state for b in BASE_IDS):
        all_dfs = {
            f"{st.session_state[f'msg_name_{b}']} - Issues": st.session_state[f"msg_{b}"][st.session_state[f"msg_{b}"]["Issue"] != "OK"]
            for b in BASE_IDS if f"msg_{b}" in st.session_state
        }
        excel_buf = build_excel(all_dfs)
        st.download_button("⬇️ Download Message Audit Report (.xlsx)",
                           excel_buf, "messages_audit_results.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    st.markdown("""
    <div style="text-align:center; padding: 32px 0 16px 0;">
        <img src="https://partingpro.com/wp-content/uploads/2024/07/partingpro-logo.png" style="height:22px; opacity:0.4;" />
        <div style="font-size:11px; color:#b0b8c8; margin-top:8px;">Aftercare Texting Audit Tool · Internal Use Only</div>
    </div>
    """, unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════════════════════
# TAB 2 — Zap Audit  (live dashboard — reads Supabase Zap Run Log)
# ════════════════════════════════════════════════════════════════════════════
# Architecture:
#   Each monitored zap → POST to a Webhooks-by-Zapier catch hook → master
#   Logger zap → POST to Supabase's REST API (table `zap_runs`). We poll
#   Supabase every 10s and render a live status board.
# ────────────────────────────────────────────────────────────────────────────

# Zap Audit — Supabase project. Set in Streamlit Cloud secrets (Settings → Secrets),
# or locally in .streamlit/secrets.toml (gitignored):
#     supabase_url = "https://xxxxxxxxxxxx.supabase.co"
#     supabase_key = "eyJ..."   # anon/public key — read-only, enforced by RLS
SUPABASE_URL = st.secrets.get("supabase_url", "")
SUPABASE_KEY = st.secrets.get("supabase_key", "")

ZAP_STATUS_META = {
    "success":   ("✅", "#1B9E6B"),
    "error":     ("❌", "#FB3D3D"),
    "halted":    ("🛑", "#E26514"),
    "held":      ("⏸",  "#e0b939"),
    "filtered":  ("🚫", "#6b7a94"),
    "delayed":   ("⏱",  "#0D6DA3"),
    "throttled": ("🐢", "#9b59b6"),
    "pending":   ("⏳", "#6b7a94"),
    "stopped":   ("💤", "#8b0000"),
}

# Zapier's own history labels (what actually lands in Supabase `status`) don't
# match the canonical bucket keys above one-for-one -- e.g. Zapier logs
# "Successful" / "Handled error" / "Safely halted" / "On hold", not
# "success" / "error" / "halted" / "held". Without this alias map, every KPI
# card except "Filtered" silently reads 0 because the lookups below are exact
# string matches against ZAP_STATUS_META's keys.
ZAP_STATUS_ALIASES = {
    "success":       "success",
    "successful":    "success",
    "error":         "error",
    "errored":       "error",
    "handled error": "error",
    "halted":        "halted",
    "safely halted": "halted",
    "held":          "held",
    "on hold":       "held",
    "filtered":      "filtered",
    "delayed":       "delayed",
    "throttled":     "throttled",
    "pending":       "pending",
    "scheduled":     "pending",
    "stopped":       "stopped",
}

def _zap_status_category(raw_status: str) -> str:
    """Map a raw (lowercased) Zapier status label to our canonical bucket key."""
    return ZAP_STATUS_ALIASES.get(raw_status, raw_status)

ZAP_WINDOW_OPTIONS = {
    "Last 1 hour":   timedelta(hours=1),
    "Last 6 hours":  timedelta(hours=6),
    "Last 24 hours": timedelta(hours=24),
    "Last 7 days":   timedelta(days=7),
    # Long windows: needed to see a zap that has gone quiet. A query zap that
    # stopped firing 9 days ago is invisible in any window shorter than this.
    "Last 30 days":  timedelta(days=30),
    "Last 60 days":  timedelta(days=60),
    "Last 90 days":  timedelta(days=90),
}

def _zap_parse_ts(s):
    """Parse Airtable ISO timestamp to timezone-aware datetime (UTC if no tz info)."""
    if not s:
        return None
    try:
        dt = datetime.fromisoformat(str(s).replace("Z", "+00:00"))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt
    except Exception:
        return None

SUPABASE_PAGE_SIZE = 1000   # PostgREST caps rows per request; page past it
SUPABASE_MAX_ROWS  = 50000  # hard stop so a runaway loop can't hang the page


@st.cache_data(ttl=30, show_spinner=False)
def fetch_zap_runs(since_iso: str = "", page_size: int = SUPABASE_PAGE_SIZE):
    """Fetch runs from Supabase (table `zap_runs`), newest first.

    `since_iso` pushes the time filter down to Postgres, so a 90-day window costs
    90 days of rows instead of the whole table. Results are paged, because the
    previous fixed `limit=500` silently truncated any window with more runs than
    that — which made a long window look emptier than it really was.
    """
    url = f"{SUPABASE_URL}/rest/v1/zap_runs"
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
    }
    records, offset = [], 0
    while offset < SUPABASE_MAX_ROWS:
        params = {
            "select": "*",
            "order": "ts.desc",
            "limit": str(page_size),
            "offset": str(offset),
        }
        if since_iso:
            params["ts"] = f"gte.{since_iso}"
        resp = requests.get(url, headers=headers, params=params, timeout=30)
        resp.raise_for_status()
        batch = resp.json()
        records.extend(batch)
        if len(batch) < page_size:
            break
        offset += len(batch)

    runs = []
    for rec in records:
        runs.append({
            "id":          rec.get("id"),
            "run_id":      rec.get("run_id", ""),
            "zap_name":    rec.get("zap_name") or "(unnamed zap)",
            "zap_id":      rec.get("zap_id", ""),
            "status":      str(rec.get("status") or "").lower(),
            "status_cat":  _zap_status_category(str(rec.get("status") or "").lower()),
            "timestamp":   rec.get("ts"),
            "step":        rec.get("step", ""),
            "error":       rec.get("error_message", ""),
            "duration_ms": rec.get("duration_ms", 0) or 0,
            "task_count":  rec.get("task_count", 0) or 0,
            "source":      rec.get("logger_source", ""),
        })
    return runs


# ── Funeral-home dimension ────────────────────────────────────────────────────
# `zap_runs` is zap-centric: it knows zap_id and zap_name but nothing about which
# funeral home a zap belongs to. The FH Contact Pull Tracker supplies that
# mapping, and is authoritative — zap names are truncated inconsistently by the
# scraper (the same zap_id appears under several name lengths), so we key on
# zap_id and fall back to parsing funeral_home_id out of the name only as a hint.
PULL_TRACKER_BASE  = "appbXFzZnhij88tnQ"          # v1 — Aftercare Texting Hub
PULL_TRACKER_TABLE = "tblRxX98yqnqrSugH"          # FH Contact Pull Tracker

# Matches "... funeral_home_id = 597", "=938", "= 1,175". Deliberately tolerant.
_FH_ID_RE = re.compile(r"funeral_home_id\s*=\s*([0-9][0-9,]*)", re.I)
# A query zap is any zap whose job is pulling contacts in for upload.
# Matched on "query data" alone, NOT the full "query data to upload in airtable":
# the scraper truncates names at varying lengths, so real rows arrive as
# "... - Query Data to upl" and a stricter pattern silently misses them.
_QUERY_ZAP_RE = re.compile(r"query\s+data", re.I)


def _extract_fh_id(zap_name: str) -> str:
    m = _FH_ID_RE.search(zap_name or "")
    return m.group(1).replace(",", "") if m else ""


# Deep link into Zapier so a 🔴 can be opened without hand-copying an ID.
#
# Only the EDITOR path is used. `zapier.com/app/editor/{id}` has been stable for years;
# Zapier has reshuffled its Zap-History URLs more than once, so a per-run deep link is
# deliberately not constructed here — a guessed run URL that 404s is worse than no link.
# Add one only after confirming the format against a real run URL.
ZAP_EDITOR_BASE = "https://zapier.com/app/editor/"


def _zap_url(zap_id) -> str:
    """Editor URL for a zap, or "" when there is nothing linkable.

    Returns "" (not a broken URL) for the three cases that legitimately have no zap:
    manual-upload homes, tracker rows with the Query Zap ID field unfilled, and the
    `fh:<id>` pseudo-IDs the orphan table uses for NULL-zap_id runs. Streamlit renders
    an empty LinkColumn cell as blank, which is the correct display for all three.
    """
    z = str(zap_id or "").strip()
    if not z or z == "—" or z.startswith("fh:"):
        return ""
    # Guard against a stray name/URL landing in the ID column — only bare numeric
    # Zap IDs produce a link.
    return ZAP_EDITOR_BASE + z if z.isdigit() else ""


def _zap_link_col(label: str = "Zap"):
    """LinkColumn showing a short label instead of the raw URL.

    `display_text` needs Streamlit >= 1.32 and requirements.txt pins no version, so
    fall back to a plain LinkColumn (then an untyped column) rather than letting a
    column-config mismatch take down the whole dashboard.
    """
    try:
        return st.column_config.LinkColumn(label, display_text="open ↗", width="small")
    except TypeError:
        try:
            return st.column_config.LinkColumn(label, width="small")
        except Exception:
            return None
    except Exception:
        return None


ZAP_LINK_COL = _zap_link_col()


# Upload outcome: "did contacts actually arrive", which run history cannot answer.
#
# Source is the `Last Contact Added` rollup on Funeral Home Information — MAX(Contact
# Created) over the linked Contact List, computed by Airtable itself. That makes the date
# live at no API cost, and reads ~4 calls per refresh across the three bases.
#
# It replaces reading `FH Upload Audit`, which stored the same date but frozen at its last
# ETL run (2026-08-21, and that ETL is not scheduled). The rollup showed several homes
# uploading as recently as 2026-08-25 that the audit table still reported as weeks idle.
FH_INFO_TABLE = "tblpf0cxsWb6Adgve"        # same table ID in all three bases
FH_INFO_BASES = {
    "v1":   "appbXFzZnhij88tnQ",
    "v1.2": "appXT2xJZ1zgll4fG",
    "v1.3": "appoDQDrqyvyPsZTY",
}
# Read by FIELD NAME, not ID: the rollup has a different field ID in each base
# (fldrEj75eC7HlSXzs / fldMr7UbV39P0Ets8 / fldiBVsmjlVuDIX6U) but one shared name.
FH_NAME_FIELD = "Funeral Home Name:"
FH_LAST_FIELD = "Last Contact Added"


def _norm_fh_name(name: str) -> str:
    """Loose match key — the two tables spell some homes slightly differently."""
    n = (name or "").lower().replace("'", "").replace("-", " ").replace(",", "").replace(".", "")
    n = n.replace("&", "and")
    return " ".join(n.split())


@st.cache_data(ttl=3600, show_spinner=False)
def fetch_last_uploads():
    """Live last-contact-added date per funeral home, keyed by normalised name.

    Reads the `Last Contact Added` rollup from Funeral Home Information in all three
    bases. A home present in more than one base keeps the LATEST date across them —
    49 homes exist in both v1 and v1.2/v1.3, and the v1 copy is usually an empty stub
    left by the migration, so taking the max avoids reporting the dead one.

    The day count is computed at render time from this date rather than stored, because
    a stored age is wrong the day after it is written.
    """
    if not AIRTABLE_READY:
        return {}
    out = {}
    for version, base in FH_INFO_BASES.items():
        url = f"https://api.airtable.com/v0/{base}/{FH_INFO_TABLE}"
        offset = None
        while True:
            params = [("pageSize", "100"),
                      ("fields[]", FH_NAME_FIELD),
                      ("fields[]", FH_LAST_FIELD)]
            if offset:
                params.append(("offset", offset))
            resp = requests.get(url, headers=HEADERS, params=params, timeout=30)
            resp.raise_for_status()
            data = resp.json()
            for rec in data.get("records", []):
                f = rec.get("fields", {})
                nm = f.get(FH_NAME_FIELD) or ""
                last = f.get(FH_LAST_FIELD) or ""
                if not nm or not last:
                    continue
                last = str(last)[:10]
                key = _norm_fh_name(nm)
                prev = out.get(key)
                if prev is None or last > prev["last"]:
                    out[key] = {"last": last, "base": version}
            offset = data.get("offset")
            if not offset:
                break
    return out


@st.cache_data(ttl=3600, show_spinner=False)
def fetch_pull_tracker():
    """One row per funeral home: name, Parting Pro ID, pull method, query zap id."""
    if not AIRTABLE_READY:
        return []
    url = f"https://api.airtable.com/v0/{PULL_TRACKER_BASE}/{PULL_TRACKER_TABLE}"
    fields = ["Funeral Home Name", "Parting Pro ID", "Pull Method",
              "Query Zap ID", "Query Zap Live", "Zapier Folder"]
    rows, offset = [], None
    while True:
        params = [("pageSize", "100")] + [("fields[]", f) for f in fields]
        if offset:
            params.append(("offset", offset))
        resp = requests.get(url, headers=HEADERS, params=params, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        for rec in data.get("records", []):
            f = rec.get("fields", {})
            rows.append({
                "name":     str(f.get("Funeral Home Name") or "").strip(),
                "pp_id":    str(f.get("Parting Pro ID") or "").strip(),
                "method":   str(f.get("Pull Method") or "").strip(),
                "zap_id":   str(f.get("Query Zap ID") or "").strip(),
                "zap_live": str(f.get("Query Zap Live") or "").strip(),
                "folder":   str(f.get("Zapier Folder") or "").strip(),
            })
        offset = data.get("offset")
        if not offset:
            break
    return rows


def render_zap_audit():
    # ── Header ────────────────────────────────────────────────────────────────
    st.markdown('''
    <div class="section-wrap">
      <div class="section-head">
        <div class="section-icon">⚡</div>
        <div class="section-head-text">
          <h3>Zap Audit — Live</h3>
          <p>Every zap run logs to Supabase; this dashboard reads it back in real time. No cookies, no polling Zapier.</p>
        </div>
      </div>
    </div>
    ''', unsafe_allow_html=True)

    # ── Controls ─────────────────────────────────────────────────────────────
    _ctrl_l, _ctrl_r = st.columns([3, 1])
    _zap_window_label = _ctrl_l.selectbox(
        "Time window",
        list(ZAP_WINDOW_OPTIONS.keys()),
        index=2,  # default: Last 24 hours
        key="zap_window_select",
    )
    if _ctrl_r.button("🔄 Refresh now", use_container_width=True, key="zap_refresh_btn"):
        fetch_zap_runs.clear()
        st.rerun()

    # ── Live dashboard — rendered inline (use Refresh button to update) ──────
    def _render_zap_dashboard():
        if not SUPABASE_URL or not SUPABASE_KEY:
            st.warning(
                "**Supabase not configured.**  \n"
                "Add these lines to your Streamlit Cloud secrets (Settings → Secrets):  \n\n"
                "`supabase_url = \"https://xxxxxxxxxxxx.supabase.co\"`  \n"
                "`supabase_key = \"eyJ...\"`  (the anon/public key)  \n\n"
                "Locally, add the same lines to `.streamlit/secrets.toml` (already gitignored)."
            )
            return
        window_delta = ZAP_WINDOW_OPTIONS[_zap_window_label]
        since_iso = (datetime.now(timezone.utc) - window_delta).isoformat()
        try:
            runs = fetch_zap_runs(since_iso=since_iso)
        except requests.HTTPError as ex:
            code = ex.response.status_code if ex.response is not None else "?"
            st.error(f"Couldn't read Zap Run Log (HTTP {code}). Check the Supabase URL/key and that the `zap_runs` table exists.")
            return
        except Exception as ex:
            st.error(f"Couldn't read Zap Run Log: {ex}")
            return

        if not runs:
            st.info(
                f"No zap runs logged in {_zap_window_label.lower()}. Once a monitored zap "
                "fires its webhook step, rows will appear here within a few seconds."
            )
            return

        # Filter to selected window
        now = datetime.now(timezone.utc)
        cutoff = now - window_delta
        in_window = []
        for r in runs:
            t = _zap_parse_ts(r["timestamp"])
            if t and t >= cutoff:
                rr = dict(r)
                rr["_t"] = t
                in_window.append(rr)

        # ── Live header ────────────────────────────────────────────────────
        st.markdown(
            f"<div style='font-size:12px; color:#6b7a94; margin: 4px 0 12px;'>"
            f"● <strong style='color: var(--pp-green);'>LIVE</strong> · "
            f"{len(in_window)} run(s) in window · "
            f"refreshed {datetime.now().strftime('%H:%M:%S')}"
            f"</div>",
            unsafe_allow_html=True,
        )

        if not in_window:
            st.caption(f"No zap runs in {_zap_window_label.lower()}. (Total in Supabase: {len(runs)})")
            return

        # ── Status summary cards ──────────────────────────────────────────
        counts = {}
        for r in in_window:
            counts[r["status_cat"]] = counts.get(r["status_cat"], 0) + 1

        # First row: Total + the 4 most common statuses
        primary_statuses = ["success", "error", "halted", "held"]
        cols = st.columns(5)
        cols[0].metric("Total runs", len(in_window))
        for i, status in enumerate(primary_statuses, 1):
            icon = ZAP_STATUS_META.get(status, ("·",))[0]
            cols[i].metric(f"{icon} {status.title()}", counts.get(status, 0))

        # Second row: other statuses, only if present
        secondary = [s for s in ("filtered", "delayed", "throttled", "pending", "stopped") if counts.get(s, 0) > 0]
        if secondary:
            cols2 = st.columns(len(secondary))
            for i, status in enumerate(secondary):
                icon = ZAP_STATUS_META.get(status, ("·",))[0]
                cols2[i].metric(f"{icon} {status.title()}", counts.get(status, 0))

        # ── Flagged section ──────────────────────────────────────────────
        flagged = {}
        for r in in_window:
            if r["status_cat"] == "error":
                flagged.setdefault(r["zap_name"], []).append(r)
        if flagged:
            st.markdown("#### 🚩 Needs attention")
            for zap_name, errors in sorted(flagged.items(), key=lambda x: -len(x[1])):
                recent = errors[0]
                step = recent["step"] or "unknown step"
                msg = (recent["error"] or "")[:150]
                detail = f"_Last error_: **{step}** — {msg}" if msg else f"_Last error_ at {step}"
                # Link off the most recent errored run — the one whose detail is shown.
                url = _zap_url(recent["zap_id"])
                if url:
                    detail += f" · [open in Zapier ↗]({url})"
                st.error(f"**{zap_name}** — {len(errors)} error(s) in window\n\n{detail}")

        # ── Activity stream (last 30 events) ────────────────────────────
        st.markdown("#### Recent activity")
        stream_rows = []
        for r in in_window[:30]:
            icon = ZAP_STATUS_META.get(r["status_cat"], ("·",))[0]
            time_str = r["_t"].astimezone().strftime("%m-%d %H:%M:%S")
            stream_rows.append({
                "Time":   time_str,
                "Status": f"{icon} {r['status']}",
                "Zap":    r["zap_name"],
                "Detail": (r["error"] or r["step"] or "")[:80],
            })
        st.dataframe(stream_rows, use_container_width=True, hide_index=True)

        # ── Per-zap aggregate table ─────────────────────────────────────
        st.markdown("#### Zap summary")
        by_zap = {}
        for r in in_window:
            zap = r["zap_name"]
            if zap not in by_zap:
                by_zap[zap] = {"name": zap, "total": 0, "last_run": None, "zap_id": ""}
                for s in ZAP_STATUS_META:
                    by_zap[zap][s] = 0
            # This table keys on zap_name, so ~9% of rows contribute no zap_id. Keep the
            # first real one seen for the link; a name whose every row is NULL stays blank.
            if not by_zap[zap]["zap_id"] and r["zap_id"]:
                by_zap[zap]["zap_id"] = str(r["zap_id"])
            by_zap[zap]["total"] += 1
            by_zap[zap][r["status_cat"]] = by_zap[zap].get(r["status_cat"], 0) + 1
            if by_zap[zap]["last_run"] is None or r["_t"] > by_zap[zap]["last_run"]:
                by_zap[zap]["last_run"] = r["_t"]

        agg_rows = []
        for z in sorted(by_zap.values(), key=lambda x: (-x.get("error", 0), -x["total"])):
            err_rate = (z.get("error", 0) / z["total"] * 100) if z["total"] else 0
            agg_rows.append({
                "Zap":      z["name"],
                "Open":     _zap_url(z["zap_id"]),
                "Total":    z["total"],
                "✅":       z.get("success", 0),
                "❌":       z.get("error", 0),
                "🛑":       z.get("halted", 0),
                "⏸":       z.get("held", 0),
                "Other":    z["total"] - sum(z.get(s, 0) for s in ("success", "error", "halted", "held")),
                "Err %":    f"{err_rate:.1f}%" if err_rate else "—",
                "Last run": z["last_run"].astimezone().strftime("%m-%d %H:%M:%S") if z["last_run"] else "—",
            })
        st.dataframe(
            agg_rows, use_container_width=True, hide_index=True,
            column_config={"Open": ZAP_LINK_COL},
        )

        # ── Funeral home coverage ────────────────────────────────────────
        st.markdown("#### Funeral home coverage")
        st.caption(
            "Joins the run log to the FH Contact Pull Tracker. Query-zap homes are "
            "judged on run history; manual-upload homes have no zap and must be "
            "judged on their last contact upload instead."
        )

        tracker = []
        uploads = {}
        if not airtable_secret_warning():
            try:
                tracker = fetch_pull_tracker()
            except requests.HTTPError as ex:
                code = ex.response.status_code if ex.response is not None else "?"
                st.error(f"Couldn't read FH Contact Pull Tracker (HTTP {code}).")
            except Exception as ex:
                st.error(f"Couldn't read FH Contact Pull Tracker: {ex}")
            try:
                uploads = fetch_last_uploads()
            except requests.HTTPError as ex:
                code = ex.response.status_code if ex.response is not None else "?"
                st.warning(f"Couldn't read last-upload dates (HTTP {code}) — "
                           f"upload columns will be blank.")
            except Exception as ex:
                st.warning(f"Couldn't read last-upload dates: {ex}")

        if tracker:
            # Aggregate by zap_id — NEVER by zap_name. The scraper truncates names
            # at inconsistent lengths, so one zap_id appears under several names.
            # ~9% of rows in zap_runs have a NULL zap_id (183 of 2,055 as of
            # 2026-08-26), and 55 of those are query zaps that still carry
            # funeral_home_id in the name. Keying on zap_id alone drops them, which
            # makes a live home read as "Silent" — a false red. So: rows WITH a
            # zap_id go in by_id, rows WITHOUT one are matched by funeral_home_id
            # instead. A row lands in exactly one index, so nothing double-counts.
            def _blank_stats():
                return {"total": 0, "success": 0, "unproductive": 0, "error": 0, "last": None}

            def _tally(d, r):
                d["total"] += 1
                cat = r["status_cat"]
                if cat == "success":
                    d["success"] += 1
                elif cat in ("filtered", "halted", "held"):
                    d["unproductive"] += 1
                elif cat == "error":
                    d["error"] += 1
                if d["last"] is None or r["_t"] > d["last"]:
                    d["last"] = r["_t"]

            def _merge(a, b):
                if not a:
                    return b
                if not b:
                    return a
                out = {k: a[k] + b[k] for k in ("total", "success", "unproductive", "error")}
                lasts = [x for x in (a["last"], b["last"]) if x]
                out["last"] = max(lasts) if lasts else None
                return out

            by_id, by_fh = {}, {}
            for r in in_window:
                zid = str(r["zap_id"] or "")
                if zid:
                    _tally(by_id.setdefault(zid, _blank_stats()), r)
                else:
                    fid = _extract_fh_id(r["zap_name"])
                    if fid:
                        _tally(by_fh.setdefault(fid, _blank_stats()), r)

            window_days = max(1, int(window_delta.total_seconds() // 86400))
            # Rank drives sort order and the "needs attention" filter.
            # 0 broken · 1 suspicious · 2 stale · 3 informational · 4 fine · 5 n/a
            RANK = {"🔴": 0, "🟠": 1, "🟡": 2, "🔵": 3, "🟢": 4, "⚪": 5}

            # CALIBRATION — measured against real data on 2026-08-26.
            #
            # 18 of 20 query zaps show zero "Successful" runs and 100% Filtered /
            # Safely halted over 30 days. Those homes are NOT behind on uploads, so
            # Filtered is plainly the normal "no new cases today" outcome for how these
            # zaps are built. (Allen Dave is the exception: no filter step, so every run
            # reads Successful.) Treating success==0 as a fault therefore paints 18 of 20
            # orange — the same uselessness as the old Needs Review field that fired on
            # 78 of 93 homes and got ignored.
            #
            # So all-filtered is INFORMATIONAL (🔵) by default. Whether contacts actually
            # arrived cannot be answered from Zapier status at all — that lives in the
            # upload counts. The checkbox below escalates it for anyone who wants to hunt.
            escalate_filtered = st.checkbox(
                "Treat all-filtered zaps as a problem",
                value=False,
                key="zap_fh_escalate_filtered",
                help="Off by default: for most of these zaps, Filtered is the normal "
                     "'no new cases' result, not a fault. Productivity has to be judged "
                     "on contacts actually arriving, not on Zapier run status.",
            )
            filtered_icon = "🟠" if escalate_filtered else "🔵"
            fh_rows = []
            for t in tracker:
                stats = _merge(
                    by_id.get(t["zap_id"]) if t["zap_id"] else None,
                    by_fh.get(t["pp_id"]) if t["pp_id"] else None,
                )
                is_query = t["method"].lower() == "query zap"

                if stats and stats["last"]:
                    days_since = (now - stats["last"]).days
                    last_str = stats["last"].astimezone().strftime("%m-%d %H:%M")
                    days_str = str(days_since)
                else:
                    days_since, last_str = None, "—"
                    days_str = f">{window_days}" if is_query else "—"

                if not is_query:
                    verdict = "⚪ Manual — judge on upload date"
                elif not t["zap_id"]:
                    verdict = "🔴 No zap ID recorded"
                elif not stats:
                    verdict = f"🔴 Silent — no runs in {window_days}d"
                # Errors are checked BEFORE "never productive": an all-error zap is
                # a different (louder) problem, and reporting it as "0 skipped" reads
                # as nonsense.
                elif stats["error"] == stats["total"]:
                    verdict = f"🔴 Every run errored ({stats['error']})"
                elif stats["success"] == 0 and stats["unproductive"] > 0:
                    verdict = (f"{filtered_icon} Runs, all filtered/halted "
                               f"({stats['unproductive']}) — check upload count")
                elif stats["success"] == 0:
                    verdict = f"🟠 No productive run in {window_days}d"
                elif days_since is not None and days_since > 3:
                    verdict = f"🟡 Stale — {days_since}d since last run"
                else:
                    verdict = "🟢 OK"

                # Upload outcome — the other half of the picture. Run history says
                # whether the mechanism fired; only this says whether contacts landed.
                up = uploads.get(_norm_fh_name(t["name"])) or {}
                last_up = up.get("last") or ""
                if last_up:
                    try:
                        d = datetime.fromisoformat(str(last_up)[:10]).date()
                        up_days = (now.date() - d).days
                        up_days_str = str(up_days)
                    except Exception:
                        up_days_str = "?"
                else:
                    up_days_str = "—"

                fh_rows.append({
                    "_rank":     RANK.get(verdict[0], 5),
                    "Funeral home": t["name"] or "(unnamed)",
                    "PP ID":     t["pp_id"] or "—",
                    "Method":    t["method"] or "(unclassified)",
                    "Zap ID":    t["zap_id"] or "—",
                    "Open":      _zap_url(t["zap_id"]),
                    "Runs":      stats["total"] if stats else 0,
                    "✅ Productive":   stats["success"] if stats else 0,
                    "🚫 Skipped":      stats["unproductive"] if stats else 0,
                    "❌ Errors":       stats["error"] if stats else 0,
                    "Last run":  last_str,
                    "Days since run": days_str,
                    "Last upload": last_up or "—",
                    "Days since upload": up_days_str,
                    "Verdict":   verdict,
                })

            fh_rows.sort(key=lambda x: (x["_rank"], x["Funeral home"].lower()))

            n_query = sum(1 for r in fh_rows if r["Method"].lower() == "query zap")
            n_bad   = sum(1 for r in fh_rows if r["_rank"] <= 1)
            n_stale = sum(1 for r in fh_rows if r["_rank"] == 2)
            n_info  = sum(1 for r in fh_rows if r["_rank"] == 3)
            n_ok    = sum(1 for r in fh_rows if r["_rank"] == 4)
            m = st.columns(6)
            m[0].metric("Homes tracked", len(fh_rows))
            m[1].metric("Query zaps", n_query)
            m[2].metric("🔴 Needs attention", n_bad)
            m[3].metric("🟡 Stale", n_stale)
            m[4].metric("🔵 All filtered", n_info)
            m[5].metric("🟢 OK", n_ok)

            # Provenance for the upload columns. Unlike the run data, this needs no
            # staleness warning: the rollup is computed by Airtable on read, so the date
            # is live. Say which homes we could not resolve instead — a blank there is
            # a name-match failure, not an idle funeral home.
            if uploads:
                unmatched = [r["Funeral home"] for r in fh_rows
                             if r["Last upload"] == "—"]
                msg = (f"**Last upload** is live — `MAX(Contact Created)` rolled up per "
                       f"funeral home across v1 / v1.2 / v1.3, computed by Airtable, no "
                       f"snapshot involved.")
                if unmatched:
                    msg += (f" {len(unmatched)} home(s) have no date: either genuinely no "
                            f"contacts yet, or the name did not match between tables — "
                            f"{', '.join(unmatched[:5])}"
                            f"{' …' if len(unmatched) > 5 else ''}.")
                st.caption(msg)
            else:
                st.caption("Upload columns unavailable — last-upload dates could not be read.")

            only_problems = st.checkbox(
                "Show only homes needing attention", value=False, key="zap_fh_only_problems"
            )
            shown = [r for r in fh_rows if r["_rank"] <= 2] if only_problems else fh_rows
            st.dataframe(
                [{k: v for k, v in r.items() if k != "_rank"} for r in shown],
                use_container_width=True, hide_index=True,
                column_config={"Open": ZAP_LINK_COL},
            )

            # Zaps present in the run log that no tracker row claims.
            claimed = {t["zap_id"] for t in tracker if t["zap_id"]}
            claimed_fh = {t["pp_id"] for t in tracker if t["pp_id"]}
            orphan_map = {}
            for r in in_window:
                zid = str(r["zap_id"] or "")
                name = r["zap_name"] or ""
                if not _QUERY_ZAP_RE.search(name):
                    continue
                if zid:
                    if zid in claimed:
                        continue
                else:
                    # No zap_id — fall back to funeral_home_id, and key the orphan
                    # on that so NULL-id rows still surface.
                    fid = _extract_fh_id(name)
                    if not fid or fid in claimed_fh:
                        continue
                    zid = f"fh:{fid}"
                # Keep the longest name seen for this zap_id — the scraper stores
                # the same zap under several truncations.
                if len(name) > len(orphan_map.get(zid, "")):
                    orphan_map[zid] = name
            orphans = sorted(
                (zid, _extract_fh_id(name), name) for zid, name in orphan_map.items()
            )
            if orphans:
                with st.expander(f"⚠️ {len(orphans)} query zap(s) running but not in the tracker"):
                    st.caption(
                        "These look like contact-pull zaps but no tracker row records their "
                        "Zap ID — either the tracker is missing a home, or the Query Zap ID "
                        "field needs filling in."
                    )
                    st.dataframe(
                        [{"Zap ID": z, "Open": _zap_url(z),
                          "funeral_home_id": f or "—", "Zap name": n}
                         for z, f, n in orphans],
                        use_container_width=True, hide_index=True,
                        column_config={"Open": ZAP_LINK_COL},
                    )

            st.caption(
                "⚠️ Freshness: this reflects whatever the Zapier history scraper has "
                "written to `zap_runs`. That scraper is a desktop-local scheduled task, so "
                "it only runs while the desktop app is up — a collection gap looks exactly "
                "like a zap going quiet. Verify a 🔴 against Zapier before acting on it."
            )

    _render_zap_dashboard()

# ── Onboarding log renderer ───────────────────────────────────────────────────
def _render_log_line(content: str):
    """Render a single log line with smart styling based on content."""
    import html as _html
    c = content.strip()
    safe = _html.escape(c)

    # Section header: ══════ or ────────
    if c.startswith("═") or c.startswith("╔") or c.startswith("╚"):
        return  # skip pure border lines — they're visual noise
    if c.startswith("╠") or c.startswith("║"):
        # Banner lines — show as bold header
        text = c.lstrip("║╠╔╚═ ").rstrip("║═ ")
        if text:
            st.markdown(
                f"<div style='background:#1e3a5f;color:#93c5fd;font-weight:700;"
                f"font-size:13px;padding:6px 12px;border-radius:4px;margin:6px 0'>"
                f"{_html.escape(text)}</div>",
                unsafe_allow_html=True
            )
        return

    # Box borders (┌ ├ └ │)
    if c.startswith("┌") or c.startswith("├") or c.startswith("└"):
        return  # skip box-drawing borders
    if c.startswith("│"):
        row = c.lstrip("│ ").rstrip("│ ")
        if row:
            st.markdown(
                f"<div style='font-size:13px;padding:2px 12px;color:#374151;"
                f"border-left:2px solid #e5e7eb;margin:1px 0'>{_html.escape(row)}</div>",
                unsafe_allow_html=True
            )
        return

    # Divider: ── or —
    if set(c.replace(" ", "")) <= {"─", "—", "-"} and len(c) > 6:
        st.markdown("<hr style='border:none;border-top:1px solid #e5e7eb;margin:8px 0'>",
                    unsafe_allow_html=True)
        return

    # List item: [1] Name — City, State (...)
    import re as _re
    list_match = _re.match(r"^\[(\d+)\]\s+(.+)$", c)
    if list_match:
        num = list_match.group(1)
        text = list_match.group(2)
        st.markdown(
            f"<div style='display:flex;gap:10px;align-items:baseline;padding:6px 10px;"
            f"background:#f8fafc;border:1px solid #e2e8f0;border-radius:6px;margin:3px 0'>"
            f"<span style='background:#3b82f6;color:white;font-size:11px;font-weight:700;"
            f"padding:2px 7px;border-radius:10px;min-width:24px;text-align:center'>{num}</span>"
            f"<span style='font-size:13px;color:#1e293b'>{_html.escape(text)}</span></div>",
            unsafe_allow_html=True
        )
        return

    # Step header: STEP X: Title
    if _re.match(r"^(STEP\s+\d|Step\s+\d)", c):
        st.markdown(
            f"<div style='background:#dbeafe;color:#1e40af;font-weight:700;"
            f"font-size:14px;padding:8px 14px;border-radius:6px;margin:8px 0'>"
            f"🔷 {safe}</div>",
            unsafe_allow_html=True
        )
        return

    # Instruction: 👉
    if c.startswith("👉") or "👉" in c[:6]:
        st.markdown(
            f"<div style='background:#fefce8;border-left:4px solid #eab308;"
            f"padding:8px 14px;border-radius:4px;margin:6px 0;font-size:13px;color:#713f12'>"
            f"{safe}</div>",
            unsafe_allow_html=True
        )
        return

    # Success: ✅
    if c.startswith("✅"):
        st.markdown(
            f"<div style='background:#f0fdf4;color:#166534;font-size:13px;"
            f"padding:5px 12px;border-radius:4px;margin:3px 0'>{safe}</div>",
            unsafe_allow_html=True
        )
        return

    # Warning: ⚠️ or ⏭️
    if c.startswith("⚠️") or c.startswith("⏭️"):
        st.markdown(
            f"<div style='background:#fff7ed;color:#9a3412;font-size:13px;"
            f"padding:5px 12px;border-radius:4px;margin:3px 0'>{safe}</div>",
            unsafe_allow_html=True
        )
        return

    # Info: ℹ️
    if c.startswith("ℹ️"):
        st.markdown(
            f"<div style='color:#475569;font-size:13px;padding:3px 12px;margin:2px 0'>"
            f"{safe}</div>",
            unsafe_allow_html=True
        )
        return

    # Link: 🔗
    if c.startswith("🔗"):
        st.markdown(
            f"<div style='background:#f0f9ff;border-left:3px solid #38bdf8;"
            f"padding:6px 12px;font-size:13px;color:#0369a1;margin:4px 0'>{safe}</div>",
            unsafe_allow_html=True
        )
        return

    # Blank / pure whitespace — skip
    if not c:
        return

    # Default: plain text
    st.markdown(
        f"<div style='font-size:13px;color:#374151;padding:2px 10px;margin:1px 0'>"
        f"{safe}</div>",
        unsafe_allow_html=True
    )


# ── Onboarding Tab ────────────────────────────────────────────────────────────
def render_onboarding():
    st.markdown("""
    <div class="section-wrap">
        <div class="section-head">
            <div class="section-icon">🚀</div>
            <div class="section-head-text">
                <h3>Funeral Home Onboarding Automation</h3>
                <p>Run automated onboarding workflows in the cloud</p>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # Initialize session state for onboarding
    if "onboarding" not in st.session_state:
        st.session_state.onboarding = None
    if "onboarding_step" not in st.session_state:
        st.session_state.onboarding_step = None
    if "onboarding_output" not in st.session_state:
        st.session_state.onboarding_output = []
    if "onboarding_input" not in st.session_state:
        st.session_state.onboarding_input = ""

    # ── How to Use ────────────────────────────────────────────────────────────────
    st.info(
        "**How to use:**  Select a step from the dropdown and click **▶️ Start Step**. "
        "Steps must be run **in order (1 → 6)** for each new funeral home. "
        "The automation will ask you questions — type your answer and press **Send**, "
        "or use the **Yes / No** buttons for confirmation prompts. "
        "Do not close or navigate away while a step is running."
    )

    st.warning(
        "⚠️ **Make sure you're working on the correct funeral home** before starting. "
        "Each step makes live changes in Airtable, Twilio, and Zapier."
    )

    # Build display list: skip Step 6 (QA/inactive), renumber Step 7 → display 6
    _display_steps = []
    _disp_num = 1
    for s in STEPS:
        if s["key"] == "6":
            continue
        _display_steps.append({**s, "display_num": str(_disp_num)})
        _disp_num += 1
    # Map display label → actual Node.js step key
    _step_option_map = {
        f"{s['emoji']} Step {s['display_num']} – {s['title']}": s["key"]
        for s in _display_steps
    }

    col_step, col_action = st.columns([3, 1])

    with col_step:
        selected_step = st.selectbox(
            "Select Onboarding Step",
            options=list(_step_option_map.keys()),
            help="Choose which onboarding step to run for the current funeral home"
        )
        step_key = _step_option_map[selected_step]  # actual key sent to Node.js

        # Show description for the selected step
        step_meta = next((s for s in STEPS if s["key"] == step_key), None)
        if step_meta:
            st.caption(f"📋 {step_meta['description']}")

    with col_action:
        st.markdown("<div style='margin-top: 30px;'></div>", unsafe_allow_html=True)
        can_start = st.session_state.onboarding is None or not st.session_state.onboarding.is_running()
        _start_help = "A step is already running — finish or cancel it first" if not can_start else None
        if st.button("▶️ Start Step", use_container_width=True, disabled=not can_start, help=_start_help):
            ob = OnboardingAutomation()
            st.session_state.onboarding = ob
            st.session_state.onboarding_step = step_key
            st.session_state.onboarding_output = []
            try:
                with st.spinner(f"Running Step {step_key}…"):
                    ob.start_step(step_key)
                    # Wait up to 15 s for the first interactive prompt
                    deadline = time.time() + 15
                    while time.time() < deadline:
                        msg = ob.get_output()
                        if msg is None:
                            time.sleep(0.15)
                            continue
                        t = msg.get("t")
                        if t == "log":
                            st.session_state.onboarding_output.append(("log", msg.get("m", "")))
                        elif t == "ask":
                            st.session_state.onboarding_output.append(("ask", msg.get("q", "")))
                            break
                        elif t == "done":
                            st.session_state.onboarding_output.append(("done", "✅ Step completed!"))
                            st.session_state.onboarding = None
                            break
                        elif t == "error":
                            st.session_state.onboarding_output.append(("error", msg.get("m", "Unknown error")))
                            st.session_state.onboarding = None
                            break
                st.rerun()
            except Exception as e:
                st.error(f"❌ Failed to start onboarding: {str(e)}")
                st.session_state.onboarding = None

    st.markdown("---")

    # Display onboarding process
    if st.session_state.onboarding and st.session_state.onboarding.is_running():
        active_step = st.session_state.get("onboarding_step", "?")
        active_meta = next((s for s in STEPS if s["key"] == active_step), None)
        active_title = active_meta["title"] if active_meta else f"Step {active_step}"
        st.markdown(
            f"<div style='background:#1e3a5f;border-left:4px solid #4a9eff;padding:10px 16px;"
            f"border-radius:4px;margin-bottom:12px;'>"
            f"<strong style='color:#4a9eff'>⚙️ Running:</strong> "
            f"<span style='color:#e0e0e0'>Step {active_step} – {active_title}</span>"
            f"</div>",
            unsafe_allow_html=True
        )

        # Current question + answer box render first, right under the "Running"
        # banner, so they're visible without scrolling past the full log below.
        input_container = st.container()
        output_container = st.container()

        # Input field for answers
        with input_container:
            history = st.session_state.onboarding_output
            last_msg = history[-1] if history else None
            if last_msg and last_msg[0] == "ask":
                # Render the context this question depends on (e.g. the record
                # details a "does this qualify?" question is asking about)
                # right here too, instead of leaving it buried in the log
                # below — everything back to the previous question, or the
                # start of the step if this is the first one.
                turn_start = 0
                for i in range(len(history) - 2, -1, -1):
                    if history[i][0] == "ask":
                        turn_start = i + 1
                        break
                for msg_type, content in history[turn_start:-1]:
                    if msg_type == "log":
                        _render_log_line(content)
                    elif msg_type == "done":
                        st.success(content)
                    elif msg_type == "error":
                        st.error(f"❌ {content}")

                question = last_msg[1]
                st.markdown(
                    f"<div style='background:#fffbeb;border-left:4px solid #f59e0b;"
                    f"padding:10px 14px;border-radius:4px;margin-bottom:8px;"
                    f"font-size:14px;color:#92400e'>"
                    f"<strong>❓ Input needed:</strong> {question}</div>",
                    unsafe_allow_html=True
                )
                is_yesno = question.rstrip().endswith("(y/n)")

                hint = ("💡 Type **y** / **n**, or use the Yes/No buttons below."
                        if is_yesno else
                        "💡 Type your answer below (e.g. a list number, or 0 to go back).")
                st.caption(hint)

                prefill = st.session_state.pop("_prefill_answer", "")
                user_input = st.text_input("Your response:", value=prefill, key="onboarding_response",
                                           placeholder="Type your answer here…")

                col_send, col_cancel = st.columns([3, 1])
                with col_send:
                    if st.button("Send ➤", use_container_width=True):
                        if user_input.strip():
                            try:
                                ob = st.session_state.onboarding
                                ob.send_answer(user_input)
                                with st.spinner("Waiting for next prompt…"):
                                    deadline = time.time() + 15
                                    while time.time() < deadline:
                                        msg = ob.get_output()
                                        if msg is None:
                                            time.sleep(0.15)
                                            continue
                                        t = msg.get("t")
                                        if t == "log":
                                            st.session_state.onboarding_output.append(("log", msg.get("m", "")))
                                        elif t == "ask":
                                            st.session_state.onboarding_output.append(("ask", msg.get("q", "")))
                                            break
                                        elif t == "done":
                                            st.session_state.onboarding_output.append(("done", "✅ Step completed!"))
                                            st.session_state.onboarding = None
                                            break
                                        elif t == "error":
                                            st.session_state.onboarding_output.append(("error", msg.get("m", "Unknown error")))
                                            st.session_state.onboarding = None
                                            break
                                st.rerun()
                            except Exception as e:
                                st.error(f"Failed to send answer: {str(e)}")
                with col_cancel:
                    if st.button("🛑 Cancel", use_container_width=True, key="btn_cancel",
                                 help="Stop the current step and discard progress"):
                        ob = st.session_state.onboarding
                        if ob:
                            ob.stop()
                        st.session_state.onboarding = None
                        st.session_state.onboarding_output = []
                        st.warning("Step cancelled.")
                        st.rerun()

                if is_yesno:
                    c1, c2 = st.columns(2)
                    with c1:
                        if st.button("✅ Yes", use_container_width=True, key="btn_yes"):
                            st.session_state["_prefill_answer"] = "y"
                            st.rerun()
                    with c2:
                        if st.button("❌ No", use_container_width=True, key="btn_no"):
                            st.session_state["_prefill_answer"] = "n"
                            st.rerun()

        # Full transcript, for scrollback — shown below the answer box instead
        # of above it, so the box doesn't keep sliding down as this grows.
        with output_container:
            st.markdown("---")
            for msg_type, content in st.session_state.onboarding_output:
                if msg_type == "log":
                    _render_log_line(content)
                elif msg_type == "ask":
                    st.markdown(
                        f"<div style='background:#fffbeb;border-left:4px solid #f59e0b;"
                        f"padding:10px 14px;border-radius:4px;margin:8px 0;"
                        f"font-size:14px;color:#92400e'>"
                        f"<strong>❓ Input needed:</strong> {content}</div>",
                        unsafe_allow_html=True
                    )
                elif msg_type == "done":
                    st.success(content)
                elif msg_type == "error":
                    st.error(f"❌ {content}")
    elif st.session_state.onboarding_output:
        st.markdown("""
        <div class="section-wrap">
            <div class="section-head">
                <div class="section-icon">✅</div>
                <div class="section-head-text">
                    <h3>Step Complete</h3>
                    <p>Process finished successfully</p>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        st.info("🎉 The onboarding step has been completed. You can start another step above or review the output below.")

        st.markdown("**Process Output:**")
        for msg_type, content in st.session_state.onboarding_output:
            if msg_type == "log":
                _render_log_line(content)
            elif msg_type == "ask":
                st.markdown(
                    f"<div style='background:#fffbeb;border-left:4px solid #f59e0b;"
                    f"padding:10px 14px;border-radius:4px;margin:8px 0;"
                    f"font-size:14px;color:#92400e'>"
                    f"<strong>❓</strong> {content}</div>",
                    unsafe_allow_html=True
                )
            elif msg_type == "done":
                st.success(content)
            elif msg_type == "error":
                st.error(f"❌ {content}")

        if st.button("Clear Output", use_container_width=True):
            st.session_state.onboarding_output = []
            st.rerun()
    else:
        st.markdown("#### 🗂️ Step Overview")
        st.caption("Run steps in order for each new funeral home. Click a step number in the dropdown above to select it, then press **▶️ Start Step**.")

        for step in _display_steps:
            st.markdown(
                f"<div style='display:flex;align-items:center;gap:14px;padding:10px 16px;"
                f"border-left:4px solid #3b82f6;background:#eff6ff;"
                f"border-radius:6px;margin-bottom:8px;'>"
                f"<span style='font-size:22px;min-width:28px'>{step['emoji']}</span>"
                f"<div>"
                f"<strong style='color:#1e40af;font-size:14px'>Step {step['display_num']}: {step['title']}</strong><br>"
                f"<span style='font-size:12px;color:#3b82f6'>{step['description']}</span>"
                f"</div></div>",
                unsafe_allow_html=True
            )

# ════════════════════════════════════════════════════════════════════════════
# TAB 4 — Task Tracker
# ════════════════════════════════════════════════════════════════════════════

# ── Task Tracker helpers ──────────────────────────────────────────────────────

def _priority_pill(p: str) -> str:
    icons = {"P1": "🔴", "P2": "🟠", "P3": "⚪"}
    cls   = {"P1": "pill-p1", "P2": "pill-p2", "P3": "pill-p3"}
    return f'<span class="{cls.get(p,"pill-p3")}">{icons.get(p,"")} {p}</span>'


def _is_overdue(task: dict) -> bool:
    if task.get("status") == "done":
        return False
    if task.get("type") != "one-off":
        return False
    due = task.get("due_date")
    if not due:
        return False
    try:
        return date.fromisoformat(str(due)) < date.today()
    except ValueError:
        return False


def _due_label(task: dict) -> str:
    due = task.get("due_date")
    if not due:
        return ""
    cls  = "overdue" if _is_overdue(task) else "due-ok"
    flag = " ⚠️" if _is_overdue(task) else ""
    return f'<span class="{cls}" style="font-size:12px;">📅 {due}{flag}</span>'


def _assignee_pills_html(task: dict) -> str:
    ids = task.get("assignee_ids") or []
    if not ids:
        return ""
    members = {m["id"]: m for m in load_members()}
    pills = []
    for mid in ids:
        m = members.get(mid)
        name = m["name"] if m else "(removed)"
        pills.append(
            f'<span style="background: var(--pp-accent-tint);color:var(--pp-accent);border-radius:10px;'
            f'padding:1px 8px;font-size:11px;margin-right:4px;">@{name}</span>'
        )
    return '<div style="margin-top:4px;">' + "".join(pills) + '</div>'


def _render_task_row(task: dict, tab_id: str = "all"):
    tid     = task["id"]
    k       = f"{tab_id}_{tid}"
    is_done = task.get("status") == "done"

    col_chk, col_info, col_type, col_due, col_edit, col_del = st.columns(
        [0.04, 0.52, 0.1, 0.18, 0.08, 0.08]
    )
    with col_chk:
        checked = st.checkbox(" ", value=is_done, key=f"chk_{k}", label_visibility="collapsed")
        if checked != is_done:
            update_task(tid, {"status": "done" if checked else "todo"})
            st.session_state.editing_task_id  = None
            st.session_state.deleting_task_id = None
            st.rerun()

    with col_info:
        title_cls = "task-title-done" if is_done else "task-title"
        desc_html = (f'<div class="task-desc">{task["description"]}</div>'
                     if task.get("description") else "")
        created_html = ""
        raw_created = task.get("created_at")
        if raw_created:
            try:
                _created_dt = datetime.fromisoformat(str(raw_created).replace("Z", "+00:00"))
                created_html = (
                    '<div style="font-size:11px;color:#9aa5b4;margin-top:2px;">'
                    f'🗓️ Created {_created_dt.strftime("%b %d, %Y")}</div>'
                )
            except Exception:
                pass
        st.markdown(
            f'<div class="{title_cls}">'
            f'{_priority_pill(task.get("priority","P3"))} {task["title"]}'
            f'</div>{desc_html}{created_html}{_assignee_pills_html(task)}',
            unsafe_allow_html=True,
        )

    with col_type:
        st.markdown(
            f'<div style="margin-top:6px;"><span class="type-badge">{task.get("type","one-off")}</span></div>',
            unsafe_allow_html=True,
        )

    with col_due:
        lbl = _due_label(task)
        if lbl:
            st.markdown(f'<div style="margin-top:8px;">{lbl}</div>', unsafe_allow_html=True)

    with col_edit:
        editing_this = st.session_state.editing_task_id == tid
        if st.button("✖️" if editing_this else "✏️", key=f"edit_btn_{k}", help="Edit"):
            st.session_state.editing_task_id  = None if editing_this else tid
            st.session_state.deleting_task_id = None
            st.rerun()

    with col_del:
        deleting_this = st.session_state.deleting_task_id == tid
        if st.button("✖️" if deleting_this else "🗑️", key=f"del_btn_{k}", help="Delete"):
            st.session_state.deleting_task_id = None if deleting_this else tid
            st.session_state.editing_task_id  = None
            st.rerun()

    # Inline edit form
    if st.session_state.editing_task_id == tid:
        _e_members = [m for m in load_members() if m.get("active")]
        _e_name_to_id = {m["name"]: m["id"] for m in _e_members}
        _e_id_to_name = {m["id"]: m["name"] for m in _e_members}
        _current_assignees = [_e_id_to_name.get(i) for i in (task.get("assignee_ids") or [])
                              if i in _e_id_to_name]
        with st.form(key=f"edit_form_{k}"):
            st.markdown("**Edit Task**")
            e_title = st.text_input("Title", value=task.get("title", ""))
            e_desc  = st.text_area("Description", value=task.get("description", ""), height=70)
            _types  = ["daily", "weekly", "monthly", "one-off"]
            e_type  = st.selectbox("Type", _types, index=_types.index(task.get("type", "one-off")))
            _pris   = ["P1", "P2", "P3"]
            e_pri   = st.selectbox("Priority", _pris, index=_pris.index(task.get("priority", "P2")))
            raw_due = task.get("due_date")
            e_due   = st.date_input("Due Date", value=date.fromisoformat(raw_due) if raw_due else None)
            e_assignees = st.multiselect("Assigned to", options=list(_e_name_to_id.keys()),
                                         default=_current_assignees)
            s_col, c_col = st.columns(2)
            with s_col: save_btn   = st.form_submit_button("💾 Save",  use_container_width=True)
            with c_col: cancel_btn = st.form_submit_button("Cancel", use_container_width=True)
        if save_btn:
            update_task(tid, {"title": e_title.strip(), "description": e_desc.strip(),
                              "type": e_type, "priority": e_pri,
                              "due_date": str(e_due) if e_due else None,
                              "assignee_ids": [_e_name_to_id[n] for n in e_assignees]})
            st.session_state.editing_task_id = None
            st.rerun()
        if cancel_btn:
            st.session_state.editing_task_id = None
            st.rerun()

    # Inline delete confirmation
    if st.session_state.deleting_task_id == tid:
        st.warning(f'Delete **"{task["title"]}"**? This cannot be undone.')
        dc, ac = st.columns(2)
        with dc:
            if st.button("🗑️ Confirm", key=f"confirm_del_{k}", use_container_width=True):
                delete_task(tid)
                st.session_state.deleting_task_id = None
                st.rerun()
        with ac:
            if st.button("Cancel", key=f"abort_del_{k}", use_container_width=True):
                st.session_state.deleting_task_id = None
                st.rerun()

    st.markdown("<hr style='margin:4px 0; border-color:#f0f2f7;'>", unsafe_allow_html=True)


def _render_task_tab(filter_type: str, all_tasks: list):
    filtered = all_tasks if filter_type == "all" else [
        t for t in all_tasks if t.get("type") == filter_type
    ]
    if not filtered:
        st.markdown(
            "<div style='padding:32px 0; text-align:center; color:#9aa5b4; font-size:14px;'>"
            "No tasks yet — add one using the sidebar form.</div>",
            unsafe_allow_html=True,
        )
        return

    pri_ord    = {"P1": 0, "P2": 1, "P3": 2}
    status_ord = {"todo": 0, "in-progress": 1, "done": 2}
    filtered   = sorted(
        filtered,
        key=lambda t: (status_ord.get(t.get("status","todo"), 0),
                       pri_ord.get(t.get("priority","P3"), 2)),
    )

    n_done    = sum(1 for t in filtered if t.get("status") == "done")
    n_overdue = sum(1 for t in filtered if _is_overdue(t))
    ov_badge  = (f' &nbsp;<span style="color: var(--pp-error);font-weight:600;">⚠️ {n_overdue} overdue</span>'
                 if n_overdue else "")
    st.markdown(
        f'<div style="font-size:13px;color:#4a5568;margin-bottom:12px;padding-bottom:8px;'
        f'border-bottom:1px solid #e4e7ef;">'
        f'<strong style="color: var(--pp-text);">{len(filtered)}</strong> tasks &nbsp;·&nbsp; '
        f'<span style="color: var(--pp-green);font-weight:600;">✅ {n_done} done</span>{ov_badge}'
        f'</div>',
        unsafe_allow_html=True,
    )
    for t in filtered:
        _render_task_row(t, tab_id=filter_type)


def render_tasks():
    tasks = load_tasks()
    today_str = date.today().isoformat()

    # ── Summary metrics ───────────────────────────────────────────────────
    active_tasks  = [t for t in tasks if t.get("status") != "done"]
    done_today    = [t for t in tasks if t.get("status") == "done"
                     and (t.get("completed_at") or "")[:10] == today_str]
    overdue_tasks = [t for t in tasks if _is_overdue(t)]
    p1_open       = [t for t in tasks if t.get("priority") == "P1" and t.get("status") != "done"]

    mc1, mc2, mc3, mc4 = st.columns(4)
    mc1.metric("Total Active", len(active_tasks),  help="All non-done tasks")
    mc2.metric("Done Today",   len(done_today),     help="Completed today")
    mc3.metric("Overdue",      len(overdue_tasks),  help="One-off tasks past due date")
    mc4.metric("P1 Items",     len(p1_open),        help="High-priority open tasks")

    st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)

    # ── Quick Capture ─────────────────────────────────────────────────────
    _qv = st.text_input("⚡ Quick add a task…", placeholder="Type and press Enter",
                        key="quick_capture", label_visibility="collapsed")
    if _qv and _qv != st.session_state.get("_last_quick", ""):
        st.session_state["_last_quick"] = _qv
        add_task({"title": _qv.strip(), "type": "one-off", "priority": "P2"})
        st.rerun()

    # ── Manage Team ───────────────────────────────────────────────────────
    with st.expander("👥 Manage Team", expanded=False):
        _all_members = load_members()
        st.caption("People available for assignment. Inactive members won't appear in the dropdowns but stay on past tasks.")

        with st.form("add_member_form", clear_on_submit=True):
            mc_a, mc_b, mc_c, mc_d = st.columns([2, 2, 2, 1])
            with mc_a: _mn = st.text_input("Name *", placeholder="Full name")
            with mc_b: _me = st.text_input("Email", placeholder="name@example.com")
            with mc_c: _mr = st.text_input("Role", placeholder="Designer, PM…")
            with mc_d:
                st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
                _madd = st.form_submit_button("➕ Add", use_container_width=True)
        if _madd:
            if _mn.strip():
                add_member({"name": _mn.strip(), "email": _me.strip(), "role": _mr.strip()})
                st.rerun()
            else:
                st.warning("Name is required.")

        if _all_members:
            for _m in _all_members:
                m_a, m_b, m_c, m_d, m_e = st.columns([2, 2, 2, 1, 1])
                m_a.markdown(f"**{_m['name']}**" + ("" if _m["active"] else " *(inactive)*"))
                m_b.markdown(f"<span style='color:#666;font-size:12px;'>{_m.get('email','—') or '—'}</span>", unsafe_allow_html=True)
                m_c.markdown(f"<span style='color:#666;font-size:12px;'>{_m.get('role','—') or '—'}</span>", unsafe_allow_html=True)
                with m_d:
                    if st.button(("Deactivate" if _m["active"] else "Activate"), key=f"toggle_{_m['id']}", use_container_width=True):
                        update_member(_m["id"], {"active": not _m["active"]})
                        st.rerun()
                with m_e:
                    if st.button("🗑️", key=f"del_member_{_m['id']}", use_container_width=True):
                        delete_member(_m["id"])
                        st.rerun()
        else:
            st.info("No team members yet. Add one above.")

    # ── Filters ────────────────────────────────────────────────────────────
    st.markdown('<div class="section-wrap">', unsafe_allow_html=True)
    st.markdown('<div class="card-title">Filters</div>', unsafe_allow_html=True)
    fc1, fc2, fc3, fc4, fc5 = st.columns([2, 2, 2, 2, 1])
    with fc1:
        priority_filter = st.multiselect("Priority", ["P1", "P2", "P3"], key="tasks_filter_priority")
    with fc2:
        status_filter = st.multiselect(
            "Status", ["todo", "done"],
            format_func=lambda s: "Done" if s == "done" else "To Do",
            key="tasks_filter_status",
        )
    with fc3:
        created_after = st.date_input("Created after", value=None, key="tasks_filter_created_after")
    with fc4:
        created_before = st.date_input("Created before", value=None, key="tasks_filter_created_before")
    with fc5:
        st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
        if st.button("Clear", use_container_width=True, key="tasks_filter_clear"):
            for k in ("tasks_filter_priority", "tasks_filter_status",
                      "tasks_filter_created_after", "tasks_filter_created_before"):
                st.session_state.pop(k, None)
            st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)

    if priority_filter:
        tasks = [t for t in tasks if t.get("priority") in priority_filter]
    if status_filter:
        tasks = [t for t in tasks if t.get("status") in status_filter]
    if created_after or created_before:
        def _created_date(t):
            try:
                return date.fromisoformat((t.get("created_at") or "")[:10])
            except Exception:
                return None

        def _in_range(t):
            d = _created_date(t)
            if d is None:
                return False
            if created_after and d < created_after:
                return False
            if created_before and d > created_before:
                return False
            return True

        tasks = [t for t in tasks if _in_range(t)]

    # ── Task Board tabs ───────────────────────────────────────────────────
    tb_all, tb_daily, tb_weekly, tb_monthly, tb_oneoff = st.tabs(
        ["All", "Daily", "Weekly", "Monthly", "One-Off"]
    )
    with tb_all:
        st.markdown('<div class="section-wrap">', unsafe_allow_html=True)
        _render_task_tab("all", tasks)
        st.markdown("</div>", unsafe_allow_html=True)
    with tb_daily:
        st.markdown('<div class="section-wrap">', unsafe_allow_html=True)
        _render_task_tab("daily", tasks)
        st.markdown("</div>", unsafe_allow_html=True)
    with tb_weekly:
        st.markdown('<div class="section-wrap">', unsafe_allow_html=True)
        _render_task_tab("weekly", tasks)
        st.markdown("</div>", unsafe_allow_html=True)
    with tb_monthly:
        st.markdown('<div class="section-wrap">', unsafe_allow_html=True)
        _render_task_tab("monthly", tasks)
        st.markdown("</div>", unsafe_allow_html=True)
    with tb_oneoff:
        st.markdown('<div class="section-wrap">', unsafe_allow_html=True)
        _render_task_tab("one-off", tasks)
        st.markdown("</div>", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════
# Dashboard — customizable task metrics/charts
# ══════════════════════════════════════════════════════════════

def render_dashboard():
    st.markdown("""
    <div class="section-wrap">
        <div class="section-head">
            <div class="section-icon">📊</div>
            <div class="section-head-text">
                <h3>Task Dashboard</h3>
                <p>Filter and explore team task activity</p>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    tasks = load_tasks()
    members = load_members()
    member_name_by_id = {m["id"]: m["name"] for m in members}

    if not tasks:
        st.info("No tasks yet — add some from the Tasks page to see them here.")
        return

    # ── Filters ────────────────────────────────────────────────────────────
    st.markdown('<div class="section-wrap">', unsafe_allow_html=True)
    f1, f2, f3, f4 = st.columns(4)
    with f1:
        type_filter = st.multiselect("Type", ["daily", "weekly", "monthly", "one-off"], key="dash_type")
    with f2:
        priority_filter = st.multiselect("Priority", ["P1", "P2", "P3"], key="dash_priority")
    with f3:
        assignee_names = sorted(m["name"] for m in members if m.get("name"))
        assignee_filter = st.multiselect("Assignee", assignee_names, key="dash_assignee")
    with f4:
        days_back = st.selectbox(
            "Created within",
            ["All time", "Last 7 days", "Last 30 days", "Last 90 days"],
            key="dash_range",
        )
    st.markdown('</div>', unsafe_allow_html=True)

    filtered = tasks
    if type_filter:
        filtered = [t for t in filtered if t.get("type") in type_filter]
    if priority_filter:
        filtered = [t for t in filtered if t.get("priority") in priority_filter]
    if assignee_filter:
        wanted_ids = {m["id"] for m in members if m.get("name") in assignee_filter}
        filtered = [t for t in filtered if wanted_ids & set(t.get("assignee_ids") or [])]
    if days_back != "All time":
        n = {"Last 7 days": 7, "Last 30 days": 30, "Last 90 days": 90}[days_back]
        cutoff = date.today() - timedelta(days=n)

        def _created_date(t):
            try:
                return date.fromisoformat((t.get("created_at") or "")[:10])
            except Exception:
                return None

        filtered = [t for t in filtered if (_created_date(t) or date.today()) >= cutoff]

    if not filtered:
        st.warning("No tasks match the current filters.")
        return

    # ── KPI cards ──────────────────────────────────────────────────────────
    total = len(filtered)
    done = [t for t in filtered if t.get("status") == "done"]
    completion_rate = (len(done) / total * 100) if total else 0
    overdue = [t for t in filtered if _is_overdue(t)]
    this_week_start = (date.today() - timedelta(days=date.today().weekday())).isoformat()
    done_this_week = [t for t in done if (t.get("completed_at") or "")[:10] >= this_week_start]

    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Total Tasks", total)
    k2.metric("Completion Rate", f"{completion_rate:.0f}%")
    k3.metric("Overdue", len(overdue))
    k4.metric("Completed This Week", len(done_this_week))

    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

    # ── Charts ─────────────────────────────────────────────────────────────
    chart_col1, chart_col2 = st.columns(2)

    with chart_col1:
        st.markdown('<div class="section-wrap">', unsafe_allow_html=True)
        st.markdown('<div class="card-title">By Priority</div>', unsafe_allow_html=True)
        pri_counts = Counter(t.get("priority", "P2") for t in filtered)
        pri_df = pd.DataFrame(
            {"Count": [pri_counts.get(p, 0) for p in ["P1", "P2", "P3"]]},
            index=["P1", "P2", "P3"],
        )
        st.bar_chart(pri_df, color="#0D6DA3", horizontal=True)
        st.markdown('</div>', unsafe_allow_html=True)

    with chart_col2:
        st.markdown('<div class="section-wrap">', unsafe_allow_html=True)
        st.markdown('<div class="card-title">By Assignee</div>', unsafe_allow_html=True)
        assignee_counts = Counter()
        for t in filtered:
            ids = t.get("assignee_ids") or []
            if not ids:
                assignee_counts["Unassigned"] += 1
            for aid in ids:
                assignee_counts[member_name_by_id.get(aid, "(removed)")] += 1
        if assignee_counts:
            assignee_df = pd.DataFrame(
                {"Count": list(assignee_counts.values())},
                index=list(assignee_counts.keys()),
            )
            st.bar_chart(assignee_df, color="#0D6DA3", horizontal=True)
        else:
            st.caption("No assignees to show.")
        st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="section-wrap">', unsafe_allow_html=True)
    st.markdown('<div class="card-title">Completions Over Time (weekly)</div>', unsafe_allow_html=True)
    weekly_counts = Counter()
    for t in done:
        try:
            d = date.fromisoformat((t.get("completed_at") or "")[:10])
        except Exception:
            continue
        week_start = d - timedelta(days=d.weekday())
        weekly_counts[week_start.isoformat()] += 1
    if weekly_counts:
        weeks_sorted = sorted(weekly_counts)
        trend_df = pd.DataFrame(
            {"Completed": [weekly_counts[k] for k in weeks_sorted]},
            index=weeks_sorted,
        )
        st.line_chart(trend_df, color="#0D6DA3")
    else:
        st.caption("No completed tasks in range yet.")
    st.markdown('</div>', unsafe_allow_html=True)


# ── Page dispatch ─────────────────────────────────────────────────────────────
# Wire the sidebar nav (see PAGE_RENDERERS init near the top) to the actual
# render functions now that they're all defined.
PAGE_RENDERERS["Onboarding"] = render_onboarding
PAGE_RENDERERS["Zap Audit"] = render_zap_audit
PAGE_RENDERERS["Tasks"] = render_tasks
PAGE_RENDERERS["Dashboard"] = render_dashboard
# Bring this back into the nav (see PAGE_RENDERERS options list up top) when ready:
#   PAGE_RENDERERS["Texting Audit"] = render_texting_audit

PAGE_RENDERERS[selected_page]()
